import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\filtered_data_references.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - REFERÊNCIA 7 - 0.004:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[2]

# inicializar listas para armazenar os dados:
feed_co2_liquid_7_004 = []
feed_co_liquid_7_004 = []
temperature_liquid_7_004 = []
pressure_liquid_7_004 = []
experimental_liquid_7_004 = []

feed_co2_liquid_7_007 = []
feed_co_liquid_7_007 = []
temperature_liquid_7_007 = []
pressure_liquid_7_007 = []
experimental_liquid_7_007 = []

feed_co2_liquid_7_0098 = []
feed_co_liquid_7_0098 = []
temperature_liquid_7_0098 = []
pressure_liquid_7_0098 = []
experimental_liquid_7_0098 = []

feed_co2_liquid_7_010107 = []
feed_co_liquid_7_010107 = []
temperature_liquid_7_010107 = []
pressure_liquid_7_010107 = []
experimental_liquid_7_010107 = []

feed_co2_liquid_7_05031 = []
feed_co_liquid_7_05031 = []
temperature_liquid_7_05031 = []
pressure_liquid_7_05031 = []
experimental_liquid_7_05031 = []

feed_co2_liquid_7_03 = []
feed_co_liquid_7_03 = []
temperature_liquid_7_03 = []
pressure_liquid_7_03 = []
experimental_liquid_7_03 = []

feed_co2_liquid_7_25254 = []
feed_co_liquid_7_25254 = []
temperature_liquid_7_25254 = []
pressure_liquid_7_25254 = []
experimental_liquid_7_25254 = []

feed_co2_liquid_7_065 = []
feed_co_liquid_7_065 = []
temperature_liquid_7_065 = []
pressure_liquid_7_065 = []
experimental_liquid_7_065 = []

feed_co2_liquid_7_50183 = []
feed_co_liquid_7_50183 = []
temperature_liquid_7_50183 = []
pressure_liquid_7_50183 = []
experimental_liquid_7_50183 = []

feed_co2_liquid_7_138 = []
feed_co_liquid_7_138 = []
temperature_liquid_7_138 = []
pressure_liquid_7_138 = []
experimental_liquid_7_138 = []

feed_co2_liquid_7_134 = []
feed_co_liquid_7_134 = []
temperature_liquid_7_134 = []
pressure_liquid_7_134 = []
experimental_liquid_7_134 = []

feed_co2_liquid_7_117 = []
feed_co_liquid_7_117 = []
temperature_liquid_7_117 = []
pressure_liquid_7_117 = []
experimental_liquid_7_117 = []

feed_co2_liquid_7_115 = []
feed_co_liquid_7_115 = []
temperature_liquid_7_115 = []
pressure_liquid_7_115 = []
experimental_liquid_7_115 = []

feed_co2_liquid_7_089 = []
feed_co_liquid_7_089 = []
temperature_liquid_7_089 = []
pressure_liquid_7_089 = []
experimental_liquid_7_089 = []

feed_co2_liquid_7_046 = []
feed_co_liquid_7_046 = []
temperature_liquid_7_046 = []
pressure_liquid_7_046 = []
experimental_liquid_7_046 = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    # verificar se o valor da coluna G (index 5) é 7
    if row[1] == 0.004:
        feed_co2_liquid_7_004.append(row[0])
        feed_co_liquid_7_004.append(row[1])
        temperature_liquid_7_004.append(row[2])
        pressure_liquid_7_004.append(row[3])
        experimental_liquid_7_004.append(row[4])
    elif row[1] == 0.007:
        feed_co2_liquid_7_007.append(row[0])
        feed_co_liquid_7_007.append(row[1])
        temperature_liquid_7_007.append(row[2])
        pressure_liquid_7_007.append(row[3])
        experimental_liquid_7_007.append(row[4])
    elif row[1] == 0.0098:
        feed_co2_liquid_7_0098.append(row[0])
        feed_co_liquid_7_0098.append(row[1])
        temperature_liquid_7_0098.append(row[2])
        pressure_liquid_7_0098.append(row[3])
        experimental_liquid_7_0098.append(row[4])
    elif row[1] == 0.010107:
        feed_co2_liquid_7_010107.append(row[0])
        feed_co_liquid_7_010107.append(row[1])
        temperature_liquid_7_010107.append(row[2])
        pressure_liquid_7_010107.append(row[3])
        experimental_liquid_7_010107.append(row[4])
    elif row[1] == 0.05031:
        feed_co2_liquid_7_05031.append(row[0])
        feed_co_liquid_7_05031.append(row[1])
        temperature_liquid_7_05031.append(row[2])
        pressure_liquid_7_05031.append(row[3])
        experimental_liquid_7_05031.append(row[4])
    elif row[1] == 0.03:
        feed_co2_liquid_7_03.append(row[0])
        feed_co_liquid_7_03.append(row[1])
        temperature_liquid_7_03.append(row[2])
        pressure_liquid_7_03.append(row[3])
        experimental_liquid_7_03.append(row[4])
    elif row[1] == 0.774:
        feed_co2_liquid_7_25254.append(row[0])
        feed_co_liquid_7_25254.append(row[1])
        temperature_liquid_7_25254.append(row[2])
        pressure_liquid_7_25254.append(row[3])
        experimental_liquid_7_25254.append(row[4])
    elif row[1] == 0.065:
        feed_co2_liquid_7_065.append(row[0])
        feed_co_liquid_7_065.append(row[1])
        temperature_liquid_7_065.append(row[2])
        pressure_liquid_7_065.append(row[3])
        experimental_liquid_7_065.append(row[4])
    elif row[1] == 0.50183:
        feed_co2_liquid_7_50183.append(row[0])
        feed_co_liquid_7_50183.append(row[1])
        temperature_liquid_7_50183.append(row[2])
        pressure_liquid_7_50183.append(row[3])
        experimental_liquid_7_50183.append(row[4])
    elif row[1] == 0.138:
        feed_co2_liquid_7_138.append(row[0])
        feed_co_liquid_7_138.append(row[1])
        temperature_liquid_7_138.append(row[2])
        pressure_liquid_7_138.append(row[3])
        experimental_liquid_7_138.append(row[4])
    elif row[1] == 0.134:
        feed_co2_liquid_7_134.append(row[0])
        feed_co_liquid_7_134.append(row[1])
        temperature_liquid_7_134.append(row[2])
        pressure_liquid_7_134.append(row[3])
        experimental_liquid_7_134.append(row[4])
    elif row[1] == 0.117:
        feed_co2_liquid_7_117.append(row[0])
        feed_co_liquid_7_117.append(row[1])
        temperature_liquid_7_117.append(row[2])
        pressure_liquid_7_117.append(row[3])
        experimental_liquid_7_117.append(row[4])
    elif row[1] == 0.115:
        feed_co2_liquid_7_115.append(row[0])
        feed_co_liquid_7_115.append(row[1])
        temperature_liquid_7_115.append(row[2])
        pressure_liquid_7_115.append(row[3])
        experimental_liquid_7_115.append(row[4])
    elif row[1] == 0.089:
        feed_co2_liquid_7_089.append(row[0])
        feed_co_liquid_7_089.append(row[1])
        temperature_liquid_7_089.append(row[2])
        pressure_liquid_7_089.append(row[3])
        experimental_liquid_7_089.append(row[4])
    elif row[1] == 0.046:
        feed_co2_liquid_7_046.append(row[0])
        feed_co_liquid_7_046.append(row[1])
        temperature_liquid_7_046.append(row[2])
        pressure_liquid_7_046.append(row[3])
        experimental_liquid_7_046.append(row[4])

# Exibindo os dados armazenados:

print("CO2 Feed (Liquid 7 - 0.004):", feed_co2_liquid_7_004)
print("CO Feed (Liquid 7 - 0.004):", feed_co_liquid_7_004)
print("Temperature (Liquid 7 - 0.004):", temperature_liquid_7_004)
print("Pressure (Liquid 7 - 0.004):", pressure_liquid_7_004)
print("Experimental (Liquid 7 - 0.004):", experimental_liquid_7_004)

print("CO2 Feed (Liquid 7 - 0.007):", feed_co2_liquid_7_007)
print("CO Feed (Liquid 7 - 0.007):", feed_co_liquid_7_007)
print("Temperature (Liquid 7 - 0.007):", temperature_liquid_7_007)
print("Pressure (Liquid 7 - 0.007):", pressure_liquid_7_007)
print("Experimental (Liquid 7 - 0.007):", experimental_liquid_7_007)

print("CO2 Feed (Liquid 7 - 0.0098):", feed_co2_liquid_7_0098)
print("CO Feed (Liquid 7 - 0.0098):", feed_co_liquid_7_0098)
print("Temperature (Liquid 7 - 0.0098):", temperature_liquid_7_0098)
print("Pressure (Liquid 7 - 0.0098):", pressure_liquid_7_0098)
print("Experimental (Liquid 7 - 0.0098):", experimental_liquid_7_0098)

print("CO2 Feed (Liquid 7 - 0.010107):", feed_co2_liquid_7_010107)
print("CO Feed (Liquid 7 - 0.010107):", feed_co_liquid_7_010107)
print("Temperature (Liquid 7 - 0.010107):", temperature_liquid_7_010107)
print("Pressure (Liquid 7 - 0.010107):", pressure_liquid_7_010107)
print("Experimental (Liquid 7 - 0.010107):", experimental_liquid_7_010107)

print("CO2 Feed (Liquid 7 - 0.05031):", feed_co2_liquid_7_05031)
print("CO Feed (Liquid 7 - 0.05031):", feed_co_liquid_7_05031)
print("Temperature (Liquid 7 - 0.05031):", temperature_liquid_7_05031)
print("Pressure (Liquid 7 - 0.05031):", pressure_liquid_7_05031)
print("Experimental (Liquid 7 - 0.05031):", experimental_liquid_7_05031)

print("CO2 Feed (Liquid 7 - 0.03):", feed_co2_liquid_7_03)
print("CO Feed (Liquid 7 - 0.03):", feed_co_liquid_7_03)
print("Temperature (Liquid 7 - 0.03):", temperature_liquid_7_03)
print("Pressure (Liquid 7 - 0.03):", pressure_liquid_7_03)
print("Experimental (Liquid 7 - 0.03):", experimental_liquid_7_03)

print("CO2 Feed (Liquid 7 - 0.774):", feed_co2_liquid_7_25254)
print("CO Feed (Liquid 7 - 0.774):", feed_co_liquid_7_25254)
print("Temperature (Liquid 7 - 0.774):", temperature_liquid_7_25254)
print("Pressure (Liquid 7 - 0.774):", pressure_liquid_7_25254)
print("Experimental (Liquid 7 - 0.774):", experimental_liquid_7_25254)

print("CO2 Feed (Liquid 7 - 0.065):", feed_co2_liquid_7_065)
print("CO Feed (Liquid 7 - 0.065):", feed_co_liquid_7_065)
print("Temperature (Liquid 7 - 0.065):", temperature_liquid_7_065)
print("Pressure (Liquid 7 - 0.065):", pressure_liquid_7_065)
print("Experimental (Liquid 7 - 0.065):", experimental_liquid_7_065)

print("CO2 Feed (Liquid 7 - 0.50183):", feed_co2_liquid_7_50183)
print("CO Feed (Liquid 7 - 0.50183):", feed_co_liquid_7_50183)
print("Temperature (Liquid 7 - 0.50183):", temperature_liquid_7_50183)
print("Pressure (Liquid 7 - 0.50183):", pressure_liquid_7_50183)
print("Experimental (Liquid 7 - 0.50183):", experimental_liquid_7_50183)

print("CO2 Feed (Liquid 7 - 0.138):", feed_co2_liquid_7_138)
print("CO Feed (Liquid 7 - 0.138):", feed_co_liquid_7_138)
print("Temperature (Liquid 7 - 0.138):", temperature_liquid_7_138)
print("Pressure (Liquid 7 - 0.138):", pressure_liquid_7_138)
print("Experimental (Liquid 7 - 0.138):", experimental_liquid_7_138)

print("CO2 Feed (Liquid 7 - 0.134):", feed_co2_liquid_7_134)
print("CO Feed (Liquid 7 - 0.134):", feed_co_liquid_7_134)
print("Temperature (Liquid 7 - 0.134):", temperature_liquid_7_134)
print("Pressure (Liquid 7 - 0.134):", pressure_liquid_7_134)
print("Experimental (Liquid 7 - 0.134):", experimental_liquid_7_134)

print("CO2 Feed (Liquid 7 - 0.117):", feed_co2_liquid_7_117)
print("CO Feed (Liquid 7 - 0.117):", feed_co_liquid_7_117)
print("Temperature (Liquid 7 - 0.117):", temperature_liquid_7_117)
print("Pressure (Liquid 7 - 0.117):", pressure_liquid_7_117)
print("Experimental (Liquid 7 - 0.117):", experimental_liquid_7_117)

print("CO2 Feed (Liquid 7 - 0.115):", feed_co2_liquid_7_115)
print("CO Feed (Liquid 7 - 0.115):", feed_co_liquid_7_115)
print("Temperature (Liquid 7 - 0.115):", temperature_liquid_7_115)
print("Pressure (Liquid 7 - 0.115):", pressure_liquid_7_115)
print("Experimental (Liquid 7 - 0.115):", experimental_liquid_7_115)

print("CO2 Feed (Liquid 7 - 0.089):", feed_co2_liquid_7_089)
print("CO Feed (Liquid 7 - 0.089):", feed_co_liquid_7_089)
print("Temperature (Liquid 7 - 0.089):", temperature_liquid_7_089)
print("Pressure (Liquid 7 - 0.089):", pressure_liquid_7_089)
print("Experimental (Liquid 7 - 0.089):", experimental_liquid_7_089)

print("CO2 Feed (Liquid 7 - 0.046):", feed_co2_liquid_7_046)
print("CO Feed (Liquid 7 - 0.046):", feed_co_liquid_7_046)
print("Temperature (Liquid 7 - 0.046):", temperature_liquid_7_046)
print("Pressure (Liquid 7 - 0.046):", pressure_liquid_7_046)
print("Experimental (Liquid 7 - 0.046):", experimental_liquid_7_046)

# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\fraction_7_liquid.xlsx'

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(0.004, (feed_co2_liquid_7_004, feed_co_liquid_7_004, temperature_liquid_7_004, pressure_liquid_7_004, experimental_liquid_7_004)),
                        (0.007, (feed_co2_liquid_7_007, feed_co_liquid_7_007, temperature_liquid_7_007, pressure_liquid_7_007, experimental_liquid_7_007)),
                        (0.0098, (feed_co2_liquid_7_0098, feed_co_liquid_7_0098, temperature_liquid_7_0098, pressure_liquid_7_0098, experimental_liquid_7_0098)),
                        (0.010107, (feed_co2_liquid_7_010107, feed_co_liquid_7_010107, temperature_liquid_7_010107, pressure_liquid_7_010107, experimental_liquid_7_010107)),
                        (0.05031, (feed_co2_liquid_7_05031, feed_co_liquid_7_05031, temperature_liquid_7_05031, pressure_liquid_7_05031, experimental_liquid_7_05031)),
                        (0.03, (feed_co2_liquid_7_03, feed_co_liquid_7_03, temperature_liquid_7_03, pressure_liquid_7_03, experimental_liquid_7_03)),
                        (0.774, (feed_co2_liquid_7_25254, feed_co_liquid_7_25254, temperature_liquid_7_25254, pressure_liquid_7_25254, experimental_liquid_7_25254)),
                        (0.065, (feed_co2_liquid_7_065, feed_co_liquid_7_065, temperature_liquid_7_065, pressure_liquid_7_065, experimental_liquid_7_065)),
                        (0.50183, (feed_co2_liquid_7_50183, feed_co_liquid_7_50183, temperature_liquid_7_50183, pressure_liquid_7_50183, experimental_liquid_7_50183)),
                        (0.138, (feed_co2_liquid_7_138, feed_co_liquid_7_138, temperature_liquid_7_138, pressure_liquid_7_138, experimental_liquid_7_138)),
                        (0.134, (feed_co2_liquid_7_134, feed_co_liquid_7_134, temperature_liquid_7_134, pressure_liquid_7_134, experimental_liquid_7_134)),
                        (0.117, (feed_co2_liquid_7_117, feed_co_liquid_7_117, temperature_liquid_7_117, pressure_liquid_7_117, experimental_liquid_7_117)),
                        (0.115, (feed_co2_liquid_7_115, feed_co_liquid_7_115, temperature_liquid_7_115, pressure_liquid_7_115, experimental_liquid_7_115)),
                        (0.089, (feed_co2_liquid_7_089, feed_co_liquid_7_089, temperature_liquid_7_089, pressure_liquid_7_089, experimental_liquid_7_089)),
                        (0.046, (feed_co2_liquid_7_046, feed_co_liquid_7_046, temperature_liquid_7_046, pressure_liquid_7_046, experimental_liquid_7_046))]:

    sheet_name = f'frac_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CO2 Liquid {ref_value}', f'Feed CO Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Experimental Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)