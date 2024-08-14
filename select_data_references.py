import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\Densidades -gásl.líquido (API só valores).xlsx'
workbook = load_workbook(file_path)

# DADOS LÍQUIDO:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[1]

# inicializar listas para armazenar os dados:
# referência 27
feed_co2_liquid_27 = []
feed_co_liquid_27 = []
temperature_liquid_27 = []
pressure_liquid_27 = []
experimental_liquid_27 = []

# referência 25
feed_co2_liquid_25 = []
feed_co_liquid_25 = []
temperature_liquid_25 = []
pressure_liquid_25 = []
experimental_liquid_25 = []

# referência 7
feed_co2_liquid_7 = []
feed_co_liquid_7 = []
temperature_liquid_7 = []
pressure_liquid_7 = []
experimental_liquid_7 = []

# iterar pelas linhas da planilha a partir da segunda linha (ignorando o cabeçalho):
for row in sheet.iter_rows(min_row=3, min_col=2, max_col=7, values_only=True):
    # verificar se o valor da coluna G (index 5) é 27
    if row[5] == 27:
        feed_co2_liquid_27.append(row[0])
        feed_co_liquid_27.append(row[1])
        temperature_liquid_27.append(row[2])
        pressure_liquid_27.append(row[3])
        experimental_liquid_27.append(row[4])
    elif row[5] == 25:
        feed_co2_liquid_25.append(row[0])
        feed_co_liquid_25.append(row[1])
        temperature_liquid_25.append(row[2])
        pressure_liquid_25.append(row[3])
        experimental_liquid_25.append(row[4])
    elif row[5] == 7:
        feed_co2_liquid_7.append(row[0])
        feed_co_liquid_7.append(row[1])
        temperature_liquid_7.append(row[2])
        pressure_liquid_7.append(row[3])
        experimental_liquid_7.append(row[4])

# Exibindo os dados armazenados:
print("CO2 Feed (Liquid 27):", feed_co2_liquid_27)
print("CO Feed (Liquid 27):", feed_co_liquid_27)
print("Temperature (Liquid 27):", temperature_liquid_27)
print("Pressure (Liquid 27):", pressure_liquid_27)
print("Experimental (Liquid 27):", experimental_liquid_27)

# Exibindo os dados armazenados:
print("CO2 Feed (Liquid 25):", feed_co2_liquid_25)
print("CO Feed (Liquid 25):", feed_co_liquid_25)
print("Temperature (Liquid 25):", temperature_liquid_25)
print("Pressure (Liquid 25):", pressure_liquid_25)
print("Experimental (Liquid 25):", experimental_liquid_25)

# Exibindo os dados armazenados:
print("CO2 Feed (Liquid 7):", feed_co2_liquid_7)
print("CO Feed (Liquid 7):", feed_co_liquid_7)
print("Temperature (Liquid 7):", temperature_liquid_7)
print("Pressure (Liquid 7):", pressure_liquid_7)
print("Experimental (Liquid 7):", experimental_liquid_7)

# DADOS GÁS:
# Selecionar a terceira aba da planilha
sheet = workbook.worksheets[2]

# Inicializar listas para armazenar os dados do gás
feed_co2_gas_27 = []
feed_co_gas_27 = []
temperature_gas_27 = []
pressure_gas_27 = []
experimental_gas_27 = []

feed_co2_gas_25 = []
feed_co_gas_25 = []
temperature_gas_25 = []
pressure_gas_25 = []
experimental_gas_25 = []

# Iterar pelas linhas da planilha a partir da terceira linha (ignorando o cabeçalho)
for row in sheet.iter_rows(min_row=3, min_col=2, max_col=7, values_only=True):
    if row[5] == 27:
        feed_co2_gas_27.append(row[0])
        feed_co_gas_27.append(row[1])
        temperature_gas_27.append(row[2])
        pressure_gas_27.append(row[3])
        experimental_gas_27.append(row[4])
    elif row[5] == 25:
        feed_co2_gas_25.append(row[0])
        feed_co_gas_25.append(row[1])
        temperature_gas_25.append(row[2])
        pressure_gas_25.append(row[3])
        experimental_gas_25.append(row[4])

# Exibindo os dados armazenados:
print("CO2 Feed (gas 27):", feed_co2_gas_27)
print("CO Feed (gas 27):", feed_co_gas_27)
print("Temperature (gas 27):", temperature_gas_27)
print("Pressure (gas 27):", pressure_gas_27)
print("Experimental (gas 27):", experimental_gas_27)

# Exibindo os dados armazenados:
print("CO2 Feed (gas 25):", feed_co2_gas_25)
print("CO Feed (gas 25):", feed_co_gas_25)
print("Temperature (gas 25):", temperature_gas_25)
print("Pressure (gas 25):", pressure_gas_25)
print("Experimental (gas 25):", experimental_gas_25)

# caminho do arquivo:
filtered_file_path = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\filtered_data_references.xlsx'

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(27, (feed_co2_liquid_27, feed_co_liquid_27, temperature_liquid_27, pressure_liquid_27, experimental_liquid_27)),
                        (25, (feed_co2_liquid_25, feed_co_liquid_25, temperature_liquid_25, pressure_liquid_25, experimental_liquid_25)),
                        (7, (feed_co2_liquid_7, feed_co_liquid_7, temperature_liquid_7, pressure_liquid_7, experimental_liquid_7))]:

    sheet_name = f'ref_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CO2 Liquid {ref_value}', f'Feed CO Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Experimental Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# adicionar os dados filtrados do gás:
for ref_value, data in [(27, (feed_co2_gas_27, feed_co_gas_27, temperature_gas_27, pressure_gas_27, experimental_gas_27)),
    (25, (feed_co2_gas_25, feed_co_gas_25, temperature_gas_25, pressure_gas_25, experimental_gas_25))]:

    sheet_name = f'ref_{ref_value}_data_gas'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CO2 Gas {ref_value}', f'Feed CO Gas {ref_value}', f'Temperature Gas {ref_value}', f'Pressure Gas {ref_value}', f'Experimental Gas {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)