import matplotlib.pyplot as plt
import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Função para plotar Histogramas:

def plot_histogram(df, title):
    df_filtered = df.dropna(axis=1, how='all')

    if not df_filtered.empty:
        df.hist(figsize=(12,10), bins=30, edgecolor='black')
        plt.suptitle(title)
        plt.tight_layout()
        plt.show()

def process_file(file_path):
    workbook= load_workbook(file_path, data_only=True)
    dados = {}
    for sheet in workbook.worksheets:
        for col in range(1, sheet.max_column + 1):
            nome_coluna = sheet.cell(row=1, column=col).value

            # Lista para armazenar os valores da coluna
            valores_coluna = [sheet.cell(row=row, column=col).value for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=col).value is not None]
            if valores_coluna: # inclui a coluna se tiver pelo menos um valor
                dados[nome_coluna] = valores_coluna
    max_len = max(len(values) for values in dados.values())
    for key, values in dados.items():
        if len(values) < max_len:
            values.extend([np.nan] * (max_len - len(values)))
    return pd.DataFrame(dados)

# Caminhos dos arquivos:
file_path_27 = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\Data Fraction Test\Fraction Liquid Test\fraction_27_liquid_test.xlsx"
file_path_25 = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\Data Fraction Test\Fraction Liquid Test\fraction_25_liquid_test.xlsx"
file_path_7 = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\Data Fraction Test\Fraction Liquid Test\fraction_7_liquid_test.xlsx"

# Processar a plotar os histogramas:
df_27 = process_file(file_path_27)
plot_histogram(df_27, 'Histogramas - Dados Líquido Test - Ref 27')

df_25 = process_file(file_path_25)
plot_histogram(df_25, 'Histogramas - Dados Líquido Test - Ref 25')

df_7 = process_file(file_path_7)
plot_histogram(df_7, 'Histogramas - Dados Líquido Test - Ref 7')

# Salvar DataFrames em um novo arquivo Excel:
outpu_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\liquid_test.xlsx"

with pd.ExcelWriter(outpu_file_path, engine='openpyxl') as writer:
    df_27.to_excel(writer, sheet_name='dados_27', index=False)
    df_25.to_excel(writer, sheet_name='dados_25', index=False)
    df_7.to_excel(writer, sheet_name='dados_7', index=False)