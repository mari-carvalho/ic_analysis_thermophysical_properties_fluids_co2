import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\data_API_tune_notune - test_treino.xlsx"
workbook = load_workbook(file_path, data_only=True)

# DADOS LÍQUIDO - REFERÊNCIA 27 - 0.004:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[1]

colunas = {}

for row in sheet.iter_rows(min_row=15, min_col=2, values_only=True):
    if any(row): # verifica se há algum valor na linha

        for idx, valor in enumerate(row): # adiciona valores a listas nas colunas correspondentes
            if idx not in colunas:
                colunas[idx] = [] # inicializa a lista para a coluna se não existir
            colunas[idx].append(valor)


for idx, dados in colunas.items():
    print(f"Coluna {idx + 2}:", dados)