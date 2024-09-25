import pandas as pd
from openpyxl import Workbook, load_workbook

file_path = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Metan Databank.csv'

df = pd.read_csv(file_path, encoding='ISO-8859-1', sep=';')

def process_value(value):

    if isinstance(value, str) and ',' in value:
        return value
    elif isinstance(value, str) and '.' in value:
        return value.replace('.', ',')
    elif isinstance(value, (float, int)):
        if value >= 1000 and value == int(value):
            return format(value/1000, '.3f').replace('.', ',')
        else:
            return format(value, '.2f').replace('.', ',')
    return value

df = df.applymap(process_value)

excel_original = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Metan Databank.xlsx'

workbook = load_workbook(excel_original)

new_workbook = Workbook()

for i, sheet in enumerate(workbook.worksheets):
    if i != 2:
        new_sheet = new_workbook.create_sheet(title=sheet.title)
        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

new_sheet = new_workbook.create_sheet(title=workbook.worksheets[2].title)

new_sheet.append(list(df.columns))

for r_idx, row in enumerate(df.itertuples(index=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        new_sheet.cell(row=r_idx, column=c_idx, value=value)

if "Sheet" in new_workbook.sheetnames:
    std = new_workbook["Sheet"]
    new_workbook.remove(std)

novo_excel = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Metan Databank NOVO.xlsx'

new_workbook.save(novo_excel)



