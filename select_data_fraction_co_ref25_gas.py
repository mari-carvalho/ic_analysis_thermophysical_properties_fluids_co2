import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\filtered_data_references.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - REFERÊNCIA 25 - 0.004:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[4]

# inicializar listas para armazenar os dados:
feed_co2_gas_25_004 = []
feed_co_gas_25_004 = []
temperature_gas_25_004 = []
pressure_gas_25_004 = []
experimental_gas_25_004 = []

feed_co2_gas_25_007 = []
feed_co_gas_25_007 = []
temperature_gas_25_007 = []
pressure_gas_25_007 = []
experimental_gas_25_007 = []

feed_co2_gas_25_0098 = []
feed_co_gas_25_0098 = []
temperature_gas_25_0098 = []
pressure_gas_25_0098 = []
experimental_gas_25_0098 = []

feed_co2_gas_25_010107 = []
feed_co_gas_25_010107 = []
temperature_gas_25_010107 = []
pressure_gas_25_010107 = []
experimental_gas_25_010107 = []

feed_co2_gas_25_05031 = []
feed_co_gas_25_05031 = []
temperature_gas_25_05031 = []
pressure_gas_25_05031 = []
experimental_gas_25_05031 = []

feed_co2_gas_25_03 = []
feed_co_gas_25_03 = []
temperature_gas_25_03 = []
pressure_gas_25_03 = []
experimental_gas_25_03 = []

feed_co2_gas_25_25254 = []
feed_co_gas_25_25254 = []
temperature_gas_25_25254 = []
pressure_gas_25_25254 = []
experimental_gas_25_25254 = []

feed_co2_gas_25_065 = []
feed_co_gas_25_065 = []
temperature_gas_25_065 = []
pressure_gas_25_065 = []
experimental_gas_25_065 = []

feed_co2_gas_25_50183 = []
feed_co_gas_25_50183 = []
temperature_gas_25_50183 = []
pressure_gas_25_50183 = []
experimental_gas_25_50183 = []

feed_co2_gas_25_138 = []
feed_co_gas_25_138 = []
temperature_gas_25_138 = []
pressure_gas_25_138 = []
experimental_gas_25_138 = []

feed_co2_gas_25_134 = []
feed_co_gas_25_134 = []
temperature_gas_25_134 = []
pressure_gas_25_134 = []
experimental_gas_25_134 = []

feed_co2_gas_25_117 = []
feed_co_gas_25_117 = []
temperature_gas_25_117 = []
pressure_gas_25_117 = []
experimental_gas_25_117 = []

feed_co2_gas_25_115 = []
feed_co_gas_25_115 = []
temperature_gas_25_115 = []
pressure_gas_25_115 = []
experimental_gas_25_115 = []

feed_co2_gas_25_089 = []
feed_co_gas_25_089 = []
temperature_gas_25_089 = []
pressure_gas_25_089 = []
experimental_gas_25_089 = []

feed_co2_gas_25_046 = []
feed_co_gas_25_046 = []
temperature_gas_25_046 = []
pressure_gas_25_046 = []
experimental_gas_25_046 = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    # verificar se o valor da coluna G (index 5) é 25
    if row[1] == 0.004:
        feed_co2_gas_25_004.append(row[0])
        feed_co_gas_25_004.append(row[1])
        temperature_gas_25_004.append(row[2])
        pressure_gas_25_004.append(row[3])
        experimental_gas_25_004.append(row[4])
    elif row[1] == 0.007:
        feed_co2_gas_25_007.append(row[0])
        feed_co_gas_25_007.append(row[1])
        temperature_gas_25_007.append(row[2])
        pressure_gas_25_007.append(row[3])
        experimental_gas_25_007.append(row[4])
    elif row[1] == 0.0098:
        feed_co2_gas_25_0098.append(row[0])
        feed_co_gas_25_0098.append(row[1])
        temperature_gas_25_0098.append(row[2])
        pressure_gas_25_0098.append(row[3])
        experimental_gas_25_0098.append(row[4])
    elif row[1] == 0.010107:
        feed_co2_gas_25_010107.append(row[0])
        feed_co_gas_25_010107.append(row[1])
        temperature_gas_25_010107.append(row[2])
        pressure_gas_25_010107.append(row[3])
        experimental_gas_25_010107.append(row[4])
    elif row[1] == 0.05031:
        feed_co2_gas_25_05031.append(row[0])
        feed_co_gas_25_05031.append(row[1])
        temperature_gas_25_05031.append(row[2])
        pressure_gas_25_05031.append(row[3])
        experimental_gas_25_05031.append(row[4])
    elif row[1] == 0.03:
        feed_co2_gas_25_03.append(row[0])
        feed_co_gas_25_03.append(row[1])
        temperature_gas_25_03.append(row[2])
        pressure_gas_25_03.append(row[3])
        experimental_gas_25_03.append(row[4])
    elif row[1] == 0.25254:
        feed_co2_gas_25_25254.append(row[0])
        feed_co_gas_25_25254.append(row[1])
        temperature_gas_25_25254.append(row[2])
        pressure_gas_25_25254.append(row[3])
        experimental_gas_25_25254.append(row[4])
    elif row[1] == 0.065:
        feed_co2_gas_25_065.append(row[0])
        feed_co_gas_25_065.append(row[1])
        temperature_gas_25_065.append(row[2])
        pressure_gas_25_065.append(row[3])
        experimental_gas_25_065.append(row[4])
    elif row[1] == 0.50183:
        feed_co2_gas_25_50183.append(row[0])
        feed_co_gas_25_50183.append(row[1])
        temperature_gas_25_50183.append(row[2])
        pressure_gas_25_50183.append(row[3])
        experimental_gas_25_50183.append(row[4])
    elif row[1] == 0.138:
        feed_co2_gas_25_138.append(row[0])
        feed_co_gas_25_138.append(row[1])
        temperature_gas_25_138.append(row[2])
        pressure_gas_25_138.append(row[3])
        experimental_gas_25_138.append(row[4])
    elif row[1] == 0.134:
        feed_co2_gas_25_134.append(row[0])
        feed_co_gas_25_134.append(row[1])
        temperature_gas_25_134.append(row[2])
        pressure_gas_25_134.append(row[3])
        experimental_gas_25_134.append(row[4])
    elif row[1] == 0.117:
        feed_co2_gas_25_117.append(row[0])
        feed_co_gas_25_117.append(row[1])
        temperature_gas_25_117.append(row[2])
        pressure_gas_25_117.append(row[3])
        experimental_gas_25_117.append(row[4])
    elif row[1] == 0.115:
        feed_co2_gas_25_115.append(row[0])
        feed_co_gas_25_115.append(row[1])
        temperature_gas_25_115.append(row[2])
        pressure_gas_25_115.append(row[3])
        experimental_gas_25_115.append(row[4])
    elif row[1] == 0.089:
        feed_co2_gas_25_089.append(row[0])
        feed_co_gas_25_089.append(row[1])
        temperature_gas_25_089.append(row[2])
        pressure_gas_25_089.append(row[3])
        experimental_gas_25_089.append(row[4])
    elif row[1] == 0.046:
        feed_co2_gas_25_046.append(row[0])
        feed_co_gas_25_046.append(row[1])
        temperature_gas_25_046.append(row[2])
        pressure_gas_25_046.append(row[3])
        experimental_gas_25_046.append(row[4])

# Exibindo os dados armazenados:

print("CO2 Feed (gas 25 - 0.004):", feed_co2_gas_25_004)
print("CO Feed (gas 25 - 0.004):", feed_co_gas_25_004)
print("Temperature (gas 25 - 0.004):", temperature_gas_25_004)
print("Pressure (gas 25 - 0.004):", pressure_gas_25_004)
print("Experimental (gas 25 - 0.004):", experimental_gas_25_004)

print("CO2 Feed (gas 25 - 0.007):", feed_co2_gas_25_007)
print("CO Feed (gas 25 - 0.007):", feed_co_gas_25_007)
print("Temperature (gas 25 - 0.007):", temperature_gas_25_007)
print("Pressure (gas 25 - 0.007):", pressure_gas_25_007)
print("Experimental (gas 25 - 0.007):", experimental_gas_25_007)

print("CO2 Feed (gas 25 - 0.0098):", feed_co2_gas_25_0098)
print("CO Feed (gas 25 - 0.0098):", feed_co_gas_25_0098)
print("Temperature (gas 25 - 0.0098):", temperature_gas_25_0098)
print("Pressure (gas 25 - 0.0098):", pressure_gas_25_0098)
print("Experimental (gas 25 - 0.0098):", experimental_gas_25_0098)

print("CO2 Feed (gas 25 - 0.010107):", feed_co2_gas_25_010107)
print("CO Feed (gas 25 - 0.010107):", feed_co_gas_25_010107)
print("Temperature (gas 25 - 0.010107):", temperature_gas_25_010107)
print("Pressure (gas 25 - 0.010107):", pressure_gas_25_010107)
print("Experimental (gas 25 - 0.010107):", experimental_gas_25_010107)

print("CO2 Feed (gas 25 - 0.05031):", feed_co2_gas_25_05031)
print("CO Feed (gas 25 - 0.05031):", feed_co_gas_25_05031)
print("Temperature (gas 25 - 0.05031):", temperature_gas_25_05031)
print("Pressure (gas 25 - 0.05031):", pressure_gas_25_05031)
print("Experimental (gas 25 - 0.05031):", experimental_gas_25_05031)

print("CO2 Feed (gas 25 - 0.03):", feed_co2_gas_25_03)
print("CO Feed (gas 25 - 0.03):", feed_co_gas_25_03)
print("Temperature (gas 25 - 0.03):", temperature_gas_25_03)
print("Pressure (gas 25 - 0.03):", pressure_gas_25_03)
print("Experimental (gas 25 - 0.03):", experimental_gas_25_03)

print("CO2 Feed (gas 25 - 0.25254):", feed_co2_gas_25_25254)
print("CO Feed (gas 25 - 0.25254):", feed_co_gas_25_25254)
print("Temperature (gas 25 - 0.25254):", temperature_gas_25_25254)
print("Pressure (gas 25 - 0.25254):", pressure_gas_25_25254)
print("Experimental (gas 25 - 0.25254):", experimental_gas_25_25254)

print("CO2 Feed (gas 25 - 0.065):", feed_co2_gas_25_065)
print("CO Feed (gas 25 - 0.065):", feed_co_gas_25_065)
print("Temperature (gas 25 - 0.065):", temperature_gas_25_065)
print("Pressure (gas 25 - 0.065):", pressure_gas_25_065)
print("Experimental (gas 25 - 0.065):", experimental_gas_25_065)

print("CO2 Feed (gas 25 - 0.50183):", feed_co2_gas_25_50183)
print("CO Feed (gas 25 - 0.50183):", feed_co_gas_25_50183)
print("Temperature (gas 25 - 0.50183):", temperature_gas_25_50183)
print("Pressure (gas 25 - 0.50183):", pressure_gas_25_50183)
print("Experimental (gas 25 - 0.50183):", experimental_gas_25_50183)

print("CO2 Feed (gas 25 - 0.138):", feed_co2_gas_25_138)
print("CO Feed (gas 25 - 0.138):", feed_co_gas_25_138)
print("Temperature (gas 25 - 0.138):", temperature_gas_25_138)
print("Pressure (gas 25 - 0.138):", pressure_gas_25_138)
print("Experimental (gas 25 - 0.138):", experimental_gas_25_138)

print("CO2 Feed (gas 25 - 0.134):", feed_co2_gas_25_134)
print("CO Feed (gas 25 - 0.134):", feed_co_gas_25_134)
print("Temperature (gas 25 - 0.134):", temperature_gas_25_134)
print("Pressure (gas 25 - 0.134):", pressure_gas_25_134)
print("Experimental (gas 25 - 0.134):", experimental_gas_25_134)

print("CO2 Feed (gas 25 - 0.117):", feed_co2_gas_25_117)
print("CO Feed (gas 25 - 0.117):", feed_co_gas_25_117)
print("Temperature (gas 25 - 0.117):", temperature_gas_25_117)
print("Pressure (gas 25 - 0.117):", pressure_gas_25_117)
print("Experimental (gas 25 - 0.117):", experimental_gas_25_117)

print("CO2 Feed (gas 25 - 0.115):", feed_co2_gas_25_115)
print("CO Feed (gas 25 - 0.115):", feed_co_gas_25_115)
print("Temperature (gas 25 - 0.115):", temperature_gas_25_115)
print("Pressure (gas 25 - 0.115):", pressure_gas_25_115)
print("Experimental (gas 25 - 0.115):", experimental_gas_25_115)

print("CO2 Feed (gas 25 - 0.089):", feed_co2_gas_25_089)
print("CO Feed (gas 25 - 0.089):", feed_co_gas_25_089)
print("Temperature (gas 25 - 0.089):", temperature_gas_25_089)
print("Pressure (gas 25 - 0.089):", pressure_gas_25_089)
print("Experimental (gas 25 - 0.089):", experimental_gas_25_089)

print("CO2 Feed (gas 25 - 0.046):", feed_co2_gas_25_046)
print("CO Feed (gas 25 - 0.046):", feed_co_gas_25_046)
print("Temperature (gas 25 - 0.046):", temperature_gas_25_046)
print("Pressure (gas 25 - 0.046):", pressure_gas_25_046)
print("Experimental (gas 25 - 0.046):", experimental_gas_25_046)

# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\fraction_25_gas.xlsx'

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(0.004, (feed_co2_gas_25_004, feed_co_gas_25_004, temperature_gas_25_004, pressure_gas_25_004, experimental_gas_25_004)),
                        (0.007, (feed_co2_gas_25_007, feed_co_gas_25_007, temperature_gas_25_007, pressure_gas_25_007, experimental_gas_25_007)),
                        (0.0098, (feed_co2_gas_25_0098, feed_co_gas_25_0098, temperature_gas_25_0098, pressure_gas_25_0098, experimental_gas_25_0098)),
                        (0.010107, (feed_co2_gas_25_010107, feed_co_gas_25_010107, temperature_gas_25_010107, pressure_gas_25_010107, experimental_gas_25_010107)),
                        (0.05031, (feed_co2_gas_25_05031, feed_co_gas_25_05031, temperature_gas_25_05031, pressure_gas_25_05031, experimental_gas_25_05031)),
                        (0.03, (feed_co2_gas_25_03, feed_co_gas_25_03, temperature_gas_25_03, pressure_gas_25_03, experimental_gas_25_03)),
                        (0.25254, (feed_co2_gas_25_25254, feed_co_gas_25_25254, temperature_gas_25_25254, pressure_gas_25_25254, experimental_gas_25_25254)),
                        (0.065, (feed_co2_gas_25_065, feed_co_gas_25_065, temperature_gas_25_065, pressure_gas_25_065, experimental_gas_25_065)),
                        (0.50183, (feed_co2_gas_25_50183, feed_co_gas_25_50183, temperature_gas_25_50183, pressure_gas_25_50183, experimental_gas_25_50183)),
                        (0.138, (feed_co2_gas_25_138, feed_co_gas_25_138, temperature_gas_25_138, pressure_gas_25_138, experimental_gas_25_138)),
                        (0.134, (feed_co2_gas_25_134, feed_co_gas_25_134, temperature_gas_25_134, pressure_gas_25_134, experimental_gas_25_134)),
                        (0.117, (feed_co2_gas_25_117, feed_co_gas_25_117, temperature_gas_25_117, pressure_gas_25_117, experimental_gas_25_117)),
                        (0.115, (feed_co2_gas_25_115, feed_co_gas_25_115, temperature_gas_25_115, pressure_gas_25_115, experimental_gas_25_115)),
                        (0.089, (feed_co2_gas_25_089, feed_co_gas_25_089, temperature_gas_25_089, pressure_gas_25_089, experimental_gas_25_089)),
                        (0.046, (feed_co2_gas_25_046, feed_co_gas_25_046, temperature_gas_25_046, pressure_gas_25_046, experimental_gas_25_046))]:

    sheet_name = f'frac_{ref_value}_data_gas'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CO2 Gas {ref_value}', f'Feed CO Gas {ref_value}', f'Temperature Gas {ref_value}', f'Pressure Gas {ref_value}', f'Experimental Gas {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)