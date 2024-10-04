import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO2+Ar\Seleção Treino_Teste\Train_Test Liquid\dados_teste_com_densidade_liquid.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
# referência 27
feed_co2_17 = []
feed_ar_17 = []
temperature_17 = []
pressure_17 = []
experimental_17 = []

# referência 25
feed_co2_18 = []
feed_ar_18 = []
temperature_18 = []
pressure_18 = []
experimental_18 = []

# referência 7
feed_co2_20 = []
feed_ar_20 = []
temperature_20 = []
pressure_20 = []
experimental_20 = []

feed_co2_22 = []
feed_ar_22 = []
temperature_22 = []
pressure_22 = []
experimental_22 = []

# iterar pelas linhas da planilha a partir da segunda linha (ignorando o cabeçalho):
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
    # verificar se o valor da coluna G (index 5) é 27
    if row[0] == 17:
        feed_co2_17.append(row[2])
        feed_ar_17.append(row[1])
        temperature_17.append(row[3])
        pressure_17.append(row[4])
        experimental_17.append(row[5])
    elif row[0] == 18:
        feed_co2_18.append(row[2])
        feed_ar_18.append(row[1])
        temperature_18.append(row[3])
        pressure_18.append(row[4])
        experimental_18.append(row[5])
    elif row[0] == 20:
        feed_co2_20.append(row[2])
        feed_ar_20.append(row[1])
        temperature_20.append(row[3])
        pressure_20.append(row[4])
        experimental_20.append(row[5])
    elif row[0] == 22:
        feed_co2_22.append(row[2])
        feed_ar_22.append(row[1])
        temperature_22.append(row[3])
        pressure_22.append(row[4])
        experimental_22.append(row[5])

# Exibindo os dados armazenados:
print("CO2 Feed (Liquid 27):", feed_co2_17)
print("CO Feed (Liquid 27):", feed_ar_17)
print("Temperature (Liquid 27):", temperature_17)
print("Pressure (Liquid 27):", pressure_17)
print("Experimental (Liquid 27):", experimental_17)

# Exibindo os dados armazenados:
print("CO2 Feed (Liquid 25):", feed_co2_18)
print("CO Feed (Liquid 25):", feed_ar_18)
print("Temperature (Liquid 25):", temperature_18)
print("Pressure (Liquid 25):", pressure_18)
print("Experimental (Liquid 25):", experimental_18)

# Exibindo os dados armazenados:
print("CO2 Feed (Liquid 7):", feed_co2_20)
print("CO Feed (Liquid 7):", feed_ar_20)
print("Temperature (Liquid 7):", temperature_20)
print("Pressure (Liquid 7):", pressure_20)
print("Experimental (Liquid 7):", experimental_20)

# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO2+Ar\Seleção Treino_Teste\Filtered References Test\References Liquid Test\filtered_data_references_liquid_test.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(17, (feed_co2_17, feed_ar_17, temperature_17, pressure_17, experimental_17)),
                        (18, (feed_co2_18, feed_ar_18, temperature_18, pressure_18, experimental_18)),
                        (20, (feed_co2_20, feed_ar_20, temperature_20, pressure_20, experimental_20)),
                        (22, (feed_co2_22, feed_ar_22, temperature_22, pressure_22, experimental_22))]:

    sheet_name = f'ref_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CO2 Liquid {ref_value}', f'Feed Ar Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Experimental Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)