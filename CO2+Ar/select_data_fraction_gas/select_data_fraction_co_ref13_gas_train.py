import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO2+Ar\Seleção Treino_Teste\Filtered References Train\References Gas Train\filtered_data_references_gas_train.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - NOVOS VALORES:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
feed_ar_13_095 = []
feed_co2_095 = []
temperature_13_095 = []
pressure_13_095 = []
experimental_13_095 = []

feed_ar_13_099 = []
feed_co2_099 = []
temperature_13_099 = []
pressure_13_099 = []
experimental_13_099 = []

feed_ar_13_050000 = []
feed_co2_050000 = []
temperature_13_050000 = []
pressure_13_050000 = []
experimental_13_050000 = []

feed_ar_13_050025 = []
feed_co2_050025 = []
temperature_13_050025 = []
pressure_13_050025 = []
experimental_13_050025 = []

feed_ar_13_024907 = []
feed_co2_024907 = []
temperature_13_024907 = []
pressure_13_024907 = []
experimental_13_024907 = []

feed_ar_13_00505 = []
feed_co2_00505 = []
temperature_13_00505 = []
pressure_13_00505 = []
experimental_13_00505 = []

feed_ar_13_02491 = []
feed_co2_02491 = []
temperature_13_02491 = []
pressure_13_02491 = []
experimental_13_02491 = []

feed_ar_13_00828 = []
feed_co2_00828 = []
temperature_13_00828 = []
pressure_13_00828 = []
experimental_13_00828 = []

feed_ar_13_01575 = []
feed_co2_01575 = []
temperature_13_01575 = []
pressure_13_01575 = []
experimental_13_01575 = []

feed_ar_13_03661 = []
feed_co2_03661 = []
temperature_13_03661 = []
pressure_13_03661 = []
experimental_13_03661 = []

feed_ar_13_04602 = []
feed_co2_04602 = []
temperature_13_04602 = []
pressure_13_04602 = []
experimental_13_04602 = []

feed_ar_13_06676 = []
feed_co2_06676 = []
temperature_13_06676 = []
pressure_13_06676 = []
experimental_13_06676 = []

feed_ar_13_07325 = []
feed_co2_07325 = []
temperature_13_07325 = []
pressure_13_07325 = []
experimental_13_07325 = []

feed_ar_13_01694 = []
feed_co2_01694 = []
temperature_13_01694 = []
pressure_13_01694 = []
experimental_13_01694 = []

feed_ar_13_0071 = []
feed_co2_0071 = []
temperature_13_0071 = []
pressure_13_0071 = []
experimental_13_0071 = []

feed_ar_13_0101 = []
feed_co2_0101 = []
temperature_13_0101 = []
pressure_13_0101 = []
experimental_13_0101 = []

feed_ar_13_0151 = []
feed_co2_0151 = []
temperature_13_0151 = []
pressure_13_0151 = []
experimental_13_0151 = []

feed_ar_13_0212 = []
feed_co2_0212 = []
temperature_13_0212 = []
pressure_13_0212 = []
experimental_13_0212 = []

feed_ar_13_0248 = []
feed_co2_0248 = []
temperature_13_0248 = []
pressure_13_0248 = []
experimental_13_0248 = []

feed_ar_13_00200 = []
feed_co2_00200 = []
temperature_13_00200 = []
pressure_13_00200 = []
experimental_13_00200 = []

feed_ar_13_00250 = []
feed_co2_00250 = []
temperature_13_00250 = []
pressure_13_00250 = []
experimental_13_00250 = []

feed_ar_13_0300 = []
feed_co2_0300 = []
temperature_13_0300 = []
pressure_13_0300 = []
experimental_13_0300 = []

feed_ar_13_025015 = []
feed_co2_025015 = []
temperature_13_025015 = []
pressure_13_025015 = []
experimental_13_025015 = []

feed_ar_13_00033 = []
feed_co2_00033 = []
temperature_13_00033 = []
pressure_13_00033 = []
experimental_13_00033 = []

feed_ar_13_00071 = []
feed_co2_00071 = []
temperature_13_00071 = []
pressure_13_00071 = []
experimental_13_00071 = []

feed_ar_13_00108 = []
feed_co2_00108 = []
temperature_13_00108 = []
pressure_13_00108 = []
experimental_13_00108 = []

feed_ar_13_00151 = []
feed_co2_00151 = []
temperature_13_00151 = []
pressure_13_00151 = []
experimental_13_00151 = []

feed_ar_13_00198 = []
feed_co2_00198 = []
temperature_13_00198 = []
pressure_13_00198 = []
experimental_13_00198 = []

feed_ar_13_00171 = []
feed_co2_00171 = []
temperature_13_00171 = []
pressure_13_00171 = []
experimental_13_00171 = []

# Adicione todas as listas necessárias...

# Preencher as listas com os dados
for row in sheet.iter_rows(min_row=2, values_only=True):  # Supondo que 'data' seja a lista de dados

    if row[1] == 0.95:
        feed_ar_13_095.append(row[1])
        feed_co2_095.append(row[0])
        temperature_13_095.append(row[2])
        pressure_13_095.append(row[3])
        experimental_13_095.append(row[4])

    elif row[1] == 0.99:
        feed_ar_13_099.append(row[1])
        feed_co2_099.append(row[0])
        temperature_13_099.append(row[2])
        pressure_13_099.append(row[3])
        experimental_13_099.append(row[4])

    elif row[1] == 0.50000:
        feed_ar_13_050000.append(row[1])
        feed_co2_050000.append(row[0])
        temperature_13_050000.append(row[2])
        pressure_13_050000.append(row[3])
        experimental_13_050000.append(row[4])

    elif row[1] == 0.50025:
        feed_ar_13_050025.append(row[1])
        feed_co2_050025.append(row[0])
        temperature_13_050025.append(row[2])
        pressure_13_050025.append(row[3])
        experimental_13_050025.append(row[4])

    elif row[1] == 0.24907:
        feed_ar_13_024907.append(row[1])
        feed_co2_024907.append(row[0])
        temperature_13_024907.append(row[2])
        pressure_13_024907.append(row[3])
        experimental_13_024907.append(row[4])

    elif row[1] == 0.0505:
        feed_ar_13_00505.append(row[1])
        feed_co2_00505.append(row[0])
        temperature_13_00505.append(row[2])
        pressure_13_00505.append(row[3])
        experimental_13_00505.append(row[4])

    elif row[1] == 0.2491:
        feed_ar_13_02491.append(row[1])
        feed_co2_02491.append(row[0])
        temperature_13_02491.append(row[2])
        pressure_13_02491.append(row[3])
        experimental_13_02491.append(row[4])

    elif row[1] == 0.0828:
        feed_ar_13_00828.append(row[1])
        feed_co2_00828.append(row[0])
        temperature_13_00828.append(row[2])
        pressure_13_00828.append(row[3])
        experimental_13_00828.append(row[4])

    elif row[1] == 0.1575:
        feed_ar_13_01575.append(row[1])
        feed_co2_01575.append(row[0])
        temperature_13_01575.append(row[2])
        pressure_13_01575.append(row[3])
        experimental_13_01575.append(row[4])

    elif row[1] == 0.3661:
        feed_ar_13_03661.append(row[1])
        feed_co2_03661.append(row[0])
        temperature_13_03661.append(row[2])
        pressure_13_03661.append(row[3])
        experimental_13_03661.append(row[4])

    elif row[0] == 0.4602:
        feed_ar_13_04602.append(row[1])
        feed_co2_04602.append(row[0])
        temperature_13_04602.append(row[2])
        pressure_13_04602.append(row[3])
        experimental_13_04602.append(row[4])

    elif row[1] == 0.6676:
        feed_ar_13_06676.append(row[1])
        feed_co2_06676.append(row[0])
        temperature_13_06676.append(row[2])
        pressure_13_06676.append(row[3])
        experimental_13_06676.append(row[4])

    elif row[1] == 0.7325:
        feed_ar_13_07325.append(row[1])
        feed_co2_07325.append(row[0])
        temperature_13_07325.append(row[2])
        pressure_13_07325.append(row[3])
        experimental_13_07325.append(row[4])

    elif row[1] == 0.1694:
        feed_ar_13_01694.append(row[1])
        feed_co2_01694.append(row[0])
        temperature_13_01694.append(row[2])
        pressure_13_01694.append(row[3])
        experimental_13_01694.append(row[4])

    elif row[1] == 0.071:
        feed_ar_13_0071.append(row[1])
        feed_co2_0071.append(row[0])
        temperature_13_0071.append(row[2])
        pressure_13_0071.append(row[3])
        experimental_13_0071.append(row[4])

    elif row[1] == 0.101:
        feed_ar_13_0101.append(row[1])
        feed_co2_0101.append(row[0])
        temperature_13_0101.append(row[2])
        pressure_13_0101.append(row[3])
        experimental_13_0101.append(row[4])

    elif row[1] == 0.151:
        feed_ar_13_0151.append(row[1])
        feed_co2_0151.append(row[0])
        temperature_13_0151.append(row[2])
        pressure_13_0151.append(row[3])
        experimental_13_0151.append(row[4])

    elif row[1] == 0.212:
        feed_ar_13_0212.append(row[1])
        feed_co2_0212.append(row[0])
        temperature_13_0212.append(row[2])
        pressure_13_0212.append(row[3])
        experimental_13_0212.append(row[4])

    elif row[1] == 0.248:
        feed_ar_13_0248.append(row[1])
        feed_co2_0248.append(row[0])
        temperature_13_0248.append(row[2])
        pressure_13_0248.append(row[3])
        experimental_13_0248.append(row[4])

    elif row[1] == 0.300:
        feed_ar_13_0300.append(row[1])
        feed_co2_0300.append(row[0])
        temperature_13_0300.append(row[2])
        pressure_13_0300.append(row[3])
        experimental_13_0300.append(row[4])

    elif row[1] == 0.25015:
        feed_ar_13_025015.append(row[1])
        feed_co2_025015.append(row[0])
        temperature_13_025015.append(row[2])
        pressure_13_025015.append(row[3])
        experimental_13_025015.append(row[4])

# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO2+Ar\Seleção Treino_Teste\Data Fraction Train\Fraction Gas Train\fraction_13_gas_train.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [
    (0.95, (feed_ar_13_095, feed_co2_095, temperature_13_095, pressure_13_095, experimental_13_095)),
    (0.99, (feed_ar_13_099, feed_co2_099, temperature_13_099, pressure_13_099, experimental_13_099)),
    (0.50000, (feed_ar_13_050000, feed_co2_050000, temperature_13_050000, pressure_13_050000, experimental_13_050000)),
    (0.50025, (feed_ar_13_050025, feed_co2_050025, temperature_13_050025, pressure_13_050025, experimental_13_050025)),
    (0.24907, (feed_ar_13_024907, feed_co2_024907, temperature_13_024907, pressure_13_024907, experimental_13_024907)),
    (0.0505, (feed_ar_13_00505, feed_co2_00505, temperature_13_00505, pressure_13_00505, experimental_13_00505)),
    (0.2491, (feed_ar_13_02491, feed_co2_02491, temperature_13_02491, pressure_13_02491, experimental_13_02491)),
    (0.0828, (feed_ar_13_00828, feed_co2_00828, temperature_13_00828, pressure_13_00828, experimental_13_00828)),
    (0.1575, (feed_ar_13_01575, feed_co2_01575, temperature_13_01575, pressure_13_01575, experimental_13_01575)),
    (0.3661, (feed_ar_13_03661, feed_co2_03661, temperature_13_03661, pressure_13_03661, experimental_13_03661)),
    (0.4602, (feed_ar_13_04602, feed_co2_04602, temperature_13_04602, pressure_13_04602, experimental_13_04602)),
    (0.6676, (feed_ar_13_06676, feed_co2_06676, temperature_13_06676, pressure_13_06676, experimental_13_06676)),
    (0.7325, (feed_ar_13_07325, feed_co2_07325, temperature_13_07325, pressure_13_07325, experimental_13_07325)),
    (0.1694, (feed_ar_13_01694, feed_co2_01694, temperature_13_01694, pressure_13_01694, experimental_13_01694)),
    (0.071, (feed_ar_13_0071, feed_co2_0071, temperature_13_0071, pressure_13_0071, experimental_13_0071)),
    (0.101, (feed_ar_13_0101, feed_co2_0101, temperature_13_0101, pressure_13_0101, experimental_13_0101)),
    (0.151, (feed_ar_13_0151, feed_co2_0151, temperature_13_0151, pressure_13_0151, experimental_13_0151)),
    (0.212, (feed_ar_13_0212, feed_co2_0212, temperature_13_0212, pressure_13_0212, experimental_13_0212)),
    (0.248, (feed_ar_13_0248, feed_co2_0248, temperature_13_0248, pressure_13_0248, experimental_13_0248)),
    (0.300, (feed_ar_13_0300, feed_co2_0300, temperature_13_0300, pressure_13_0300, experimental_13_0300)),
    (0.25015, (feed_ar_13_025015, feed_co2_025015, temperature_13_025015, pressure_13_025015, experimental_13_025015))]:

    sheet_name = f'frac_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CH4 Liquid {ref_value}',  f'Feed CO2 Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Rho Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)