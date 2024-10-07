import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO2+Ar\Seleção Treino_Teste\Filtered References Test\References Liquid Test\filtered_data_references_liquid_test.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - NOVOS VALORES:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[1]

# inicializar listas para armazenar os dados:
feed_ar_18_024907 = []
feed_co2_024907 = []
temperature_18_024907 = []
pressure_18_024907 = []
experimental_18_024907 = []

feed_ar_18_00308 = []
feed_co2_00308 = []
temperature_18_00308 = []
pressure_18_00308 = []
experimental_18_00308 = []

feed_ar_18_01694 = []
feed_co2_01694 = []
temperature_18_01694 = []
pressure_18_01694 = []
experimental_18_01694 = []

feed_ar_18_0207 = []
feed_co2_0207 = []
temperature_18_0207 = []
pressure_18_0207 = []
experimental_18_0207 = []

feed_ar_18_0248 = []
feed_co2_0248 = []
temperature_18_0248 = []
pressure_18_0248 = []
experimental_18_0248 = []

feed_ar_18_0060 = []
feed_co2_0060 = []
temperature_18_0060 = []
pressure_18_0060 = []
experimental_18_0060 = []

feed_ar_18_0100 = []
feed_co2_0100 = []
temperature_18_0100 = []
pressure_18_0100 = []
experimental_18_0100 = []

feed_ar_18_0134 = []
feed_co2_0134 = []
temperature_18_0134 = []
pressure_18_0134 = []
experimental_18_0134 = []

feed_ar_18_0167 = []
feed_co2_0167 = []
temperature_18_0167 = []
pressure_18_0167 = []
experimental_18_0167 = []

feed_ar_18_050025 = []
feed_co2_050025 = []
temperature_18_050025 = []
pressure_18_050025 = []
experimental_18_050025 = []

feed_ar_18_025015 = []
feed_co2_025015 = []
temperature_18_025015 = []
pressure_18_025015 = []
experimental_18_025015 = []

feed_ar_18_00009 = []
feed_co2_00009 = []
temperature_18_00009 = []
pressure_18_00009 = []
experimental_18_00009 = []

feed_ar_18_00014 = []
feed_co2_00014 = []
temperature_18_00014 = []
pressure_18_00014 = []
experimental_18_00014 = []

feed_ar_18_00024 = []
feed_co2_00024 = []
temperature_18_00024 = []
pressure_18_00024 = []
experimental_18_00024 = []

feed_ar_18_00041 = []
feed_co2_00041 = []
temperature_18_00041 = []
pressure_18_00041 = []
experimental_18_00041 = []

feed_ar_18_00063 = []
feed_co2_00063 = []
temperature_18_00063 = []
pressure_18_00063 = []
experimental_18_00063 = []

feed_ar_18_00084 = []
feed_co2_00084 = []
temperature_18_00084 = []
pressure_18_00084 = []
experimental_18_00084 = []

feed_ar_18_00142 = []
feed_co2_00142 = []
temperature_18_00142 = []
pressure_18_00142 = []
experimental_18_00142 = []

feed_ar_18_00020 = []
feed_co2_00020 = []
temperature_18_00020 = []
pressure_18_00020 = []
experimental_18_00020 = []

feed_ar_18_00200 = []
feed_co2_00200 = []
temperature_18_00200 = []
pressure_18_00200 = []
experimental_18_00200 = []

feed_ar_18_00250 = []
feed_co2_00250 = []
temperature_18_00250 = []
pressure_18_00250 = []
experimental_18_00250 = []

feed_ar_18_00011 = []
feed_co2_00011 = []
temperature_18_00011 = []
pressure_18_00011 = []
experimental_18_00011 = []

feed_ar_18_00007 = []
feed_co2_00007 = []
temperature_18_00007 = []
pressure_18_00007 = []
experimental_18_00007 = []

feed_ar_18_00033 = []
feed_co2_00033 = []
temperature_18_00033 = []
pressure_18_00033 = []
experimental_18_00033 = []

feed_ar_18_00071 = []
feed_co2_00071 = []
temperature_18_00071 = []
pressure_18_00071 = []
experimental_18_00071 = []

feed_ar_18_00108 = []
feed_co2_00108 = []
temperature_18_00108 = []
pressure_18_00108 = []
experimental_18_00108 = []

feed_ar_18_00151 = []
feed_co2_00151 = []
temperature_18_00151 = []
pressure_18_00151 = []
experimental_18_00151 = []

feed_ar_18_00198 = []
feed_co2_00198 = []
temperature_18_00198 = []
pressure_18_00198 = []
experimental_18_00198 = []

feed_ar_18_00171 = []
feed_co2_00171 = []
temperature_18_00171 = []
pressure_18_00171 = []
experimental_18_00171 = []

# Adicione todas as listas necessárias...

# Preencher as listas com os dados
for row in sheet.iter_rows(min_row=2, values_only=True):  # Supondo que 'data' seja a lista de dados

    if row[1] == 0.24907:
        feed_ar_18_024907.append(row[1])
        feed_co2_024907.append(row[0])
        temperature_18_024907.append(row[2])
        pressure_18_024907.append(row[3])
        experimental_18_024907.append(row[4])

    elif row[1] == 0.0308:
        feed_ar_18_00308.append(row[1])
        feed_co2_00308.append(row[0])
        temperature_18_00308.append(row[2])
        pressure_18_00308.append(row[3])
        experimental_18_00308.append(row[4])

    elif row[1] == 0.1694:
        feed_ar_18_01694.append(row[1])
        feed_co2_01694.append(row[0])
        temperature_18_01694.append(row[2])
        pressure_18_01694.append(row[3])
        experimental_18_01694.append(row[4])

    elif row[1] == 0.207:
        feed_ar_18_0207.append(row[1])
        feed_co2_0207.append(row[0])
        temperature_18_0207.append(row[2])
        pressure_18_0207.append(row[3])
        experimental_18_0207.append(row[4])

    elif row[1] == 0.248:
        feed_ar_18_0248.append(row[1])
        feed_co2_0248.append(row[0])
        temperature_18_0248.append(row[2])
        pressure_18_0248.append(row[3])
        experimental_18_0248.append(row[4])

    elif row[1] == 0.060:
        feed_ar_18_0060.append(row[1])
        feed_co2_0060.append(row[0])
        temperature_18_0060.append(row[2])
        pressure_18_0060.append(row[3])
        experimental_18_0060.append(row[4])

    elif row[1] == 0.100:
        feed_ar_18_0100.append(row[1])
        feed_co2_0100.append(row[0])
        temperature_18_0100.append(row[2])
        pressure_18_0100.append(row[3])
        experimental_18_0100.append(row[4])

    elif row[1] == 0.134:
        feed_ar_18_0134.append(row[1])
        feed_co2_0134.append(row[0])
        temperature_18_0134.append(row[2])
        pressure_18_0134.append(row[3])
        experimental_18_0134.append(row[4])

    elif row[1] == 0.167:
        feed_ar_18_0167.append(row[1])
        feed_co2_0167.append(row[0])
        temperature_18_0167.append(row[2])
        pressure_18_0167.append(row[3])
        experimental_18_0167.append(row[4])

    elif row[1] == 0.50025:
        feed_ar_18_050025.append(row[1])
        feed_co2_050025.append(row[0])
        temperature_18_050025.append(row[2])
        pressure_18_050025.append(row[3])
        experimental_18_050025.append(row[4])

    elif row[1] == 0.25015:
        feed_ar_18_025015.append(row[1])
        feed_co2_025015.append(row[0])
        temperature_18_025015.append(row[2])
        pressure_18_025015.append(row[3])
        experimental_18_025015.append(row[4])


# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO2+Ar\Seleção Treino_Teste\Data Fraction Test\Fraction Liquid Test\fraction_18_liquid_test.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [
    (0.24907, (feed_ar_18_024907, feed_co2_024907, temperature_18_024907, pressure_18_024907, experimental_18_024907)),
    (0.0308, (feed_ar_18_00308, feed_co2_00308, temperature_18_00308, pressure_18_00308, experimental_18_00308)),
    (0.1694, (feed_ar_18_01694, feed_co2_01694, temperature_18_01694, pressure_18_01694, experimental_18_01694)),
    (0.207, (feed_ar_18_0207, feed_co2_0207, temperature_18_0207, pressure_18_0207, experimental_18_0207)),
    (0.248, (feed_ar_18_0248, feed_co2_0248, temperature_18_0248, pressure_18_0248, experimental_18_0248)),
    (0.060, (feed_ar_18_0060, feed_co2_0060, temperature_18_0060, pressure_18_0060, experimental_18_0060)),
    (0.100, (feed_ar_18_0100, feed_co2_0100, temperature_18_0100, pressure_18_0100, experimental_18_0100)),
    (0.134, (feed_ar_18_0134, feed_co2_0134, temperature_18_0134, pressure_18_0134, experimental_18_0134)),
    (0.167, (feed_ar_18_0167, feed_co2_0167, temperature_18_0167, pressure_18_0167, experimental_18_0167)),
    (0.50025, (feed_ar_18_050025, feed_co2_050025, temperature_18_050025, pressure_18_050025, experimental_18_050025)),
    (0.25015, (feed_ar_18_025015, feed_co2_025015, temperature_18_025015, pressure_18_025015, experimental_18_025015))]:

    sheet_name = f'frac_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CH4 Liquid {ref_value}',  f'Feed CO2 Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Rho Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)