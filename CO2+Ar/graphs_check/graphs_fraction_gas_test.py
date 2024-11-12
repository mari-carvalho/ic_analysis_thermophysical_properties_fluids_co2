import matplotlib.pyplot as plt
import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Função para plotar Histogramas:

def plot_histogram(df, title):
    df_filtered = df.dropna(axis=1, how='all')

    if not df_filtered.empty:
        df.hist(figsize=(12,10), bins=30, edgecolor='pink')
        plt.suptitle(title)

        for idx, ax in enumerate(plt.gcf().axes):
            nome_coluna = df_filtered.columns[idx]
            palavras = nome_coluna.split()

            if idx < 2:
                novo_rotulo = ' '.join(palavras[:2])
            else:
                novo_rotulo = palavras[0]

            ax.set_xlabel(novo_rotulo)
            ax.set_ylabel('Frequency')



        plt.tight_layout()
        plt.show()

def process_file(file_path):
    workbook= load_workbook(file_path, data_only=True) # carreamento da planilha; data_only=True - indica que somente os valores das células serão lidos, desconsiderando fórmulas
    dados = {} # criando um dicionário para armazenar os dados de cada coluna
    for sheet in workbook.worksheets: # itera sobre todas as abas da planilha
        for col in range(1, sheet.max_column + 1): # para cada aba, itera sobre cada coluna
            nome_coluna = sheet.cell(row=1, column=col).value # o nome da coluna é obtido a partir da primeira linha da coluna da vez

            # Lista para armazenar os valores da coluna
            valores_coluna = [sheet.cell(row=row, column=col).value for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=col).value is not None] # uma lista com todos os valores da coluna da vez a partir da segunda linha, ignorando células vazias
            if valores_coluna: # inclui a coluna se tiver pelo menos um valor
                dados[nome_coluna] = valores_coluna # adiciona a lista anterior ao dicionário se a mesma não estiver vazia, usando o nome da coluna como chave
    max_len = max(len(values) for values in dados.values()) # dados.values retorna todas as listas de valores armazenadas no dicionário; len(values) calcula o número de elementos; max determina o maior valor entre os tamanhos das litas
    for key, values in dados.items(): # itera sobre cada entrada no dicionário; dados.key indica o nome da coluna; values é a lista de valores correspondente
        if len(values) < max_len: # verifica se o número de valores é menor que a lista de maior tamanho
            values.extend([np.nan] * (max_len - len(values))) # max_len - len(values) calcula quantos valores precisam ser adicionados para a lista alcançar o valor máximo de elementos; [np.nan] cria uma lista com NaN repetido quantas vezes forem necessários para a lista ter o número máximo de elementos; values.extend adiciona NaN ao final da lista values para a lista ter o número máximo de elementos
    return pd.DataFrame(dados)

# Caminhos dos arquivos:
file_path_27 = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO+CO2\Seleção Treino_Teste\Data Fraction Test\Fraction Gas Test\fraction_27_gas_test.xlsx"
file_path_25 = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO+CO2\Seleção Treino_Teste\Data Fraction Test\Fraction Gas Test\fraction_25_gas_test.xlsx"
file_path_7 = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO+CO2\Seleção Treino_Teste\Data Fraction Test\Fraction Gas Test\fraction_7_gas_test.xlsx"

# Processar a plotar os histogramas:
df_27 = process_file(file_path_27)
plot_histogram(df_27, 'Histograms - Data Gas Test - Ref 27')

df_25 = process_file(file_path_25)
plot_histogram(df_25, 'Histograms - Data Gas Test - Ref 25')

df_7 = process_file(file_path_7)
plot_histogram(df_7, 'Histograms - Data Gas Test - Ref 7')

# Salvar DataFrames em um novo arquivo Excel:
outpu_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\gas_test.xlsx"

with pd.ExcelWriter(outpu_file_path, engine='openpyxl') as writer:
    df_27.to_excel(writer, sheet_name='dados_27', index=False)
    df_25.to_excel(writer, sheet_name='dados_25', index=False)
    df_7.to_excel(writer, sheet_name='dados_7', index=False)