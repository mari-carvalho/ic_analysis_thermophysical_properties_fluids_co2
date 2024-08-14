import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook


files = [
    r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\fraction_27_liquid.xlsx",
    r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\fraction_25_liquid.xlsx",
    r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\fraction_7_liquid.xlsx",
    r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\fraction_27_gas.xlsx",
    r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\fraction_25_gas.xlsx"
]

wb_unido = Workbook()
ws_unido = wb_unido.active
ws_unido.title = "Unificado"

linha_atual = 1

for file in files:
    wb = load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        ws_unido.append(row)
        linha_atual += 1


wb_unido.save("tabela_unificada.xlsx")