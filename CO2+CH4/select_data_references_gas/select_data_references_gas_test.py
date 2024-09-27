import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Train_Test Gas\dados_teste_com_densidade_gas.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
# referência 36
feed_ch4_36 = []
v_36 = []
temperature_36 = []
pressure_36 = []
experimental_36 = []


# iterar pelas linhas da planilha a partir da segunda linha (ignorando o cabeçalho):
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
    # verificar se o valor da coluna G (index 5) é 36
    if row[0] == 36:
        feed_ch4_36.append(row[1])
        temperature_36.append(row[2])
        pressure_36.append(row[3])
        v_36.append(row[4])
        experimental_36.append(row[5])

# Exibindo os dados armazenados:
print("CH4 Feed (gas 36):", feed_ch4_36)
print("V (gas 36):", v_36)
print("Temperature (gas 36):", temperature_36)
print("Pressure (gas 36):", pressure_36)
print("Experimental (gas 36):", experimental_36)


# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Filtered References Test\References Gas Test\filtered_data_references_gas_test.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(36, (feed_ch4_36, temperature_36, pressure_36, v_36, experimental_36))]:

    sheet_name = f'ref_{ref_value}_data_gas'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CH4 gas {ref_value}', f'Temperature Gas {ref_value}', f'Pressure Gas {ref_value}', f'V Gas {ref_value}', f'Rho Gas {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)