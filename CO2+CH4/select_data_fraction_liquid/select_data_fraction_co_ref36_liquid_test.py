import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Filtered References Test\References Liquid Test\filtered_data_references_liquid_test.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - REFERÊNCIA 27 - 0.004:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
feed_ch4_36_0043 = []
v_36_0043 = []
temperature_36_0043 = []
pressure_36_0043 = []
experimental_36_0043 = []

feed_ch4_36_0091 = []
v_36_0091 = []
temperature_36_0091 = []
pressure_36_0091 = []
experimental_36_0091 = []

feed_ch4_36_0116 = []
v_36_0116 = []
temperature_36_0116 = []
pressure_36_0116 = []
experimental_36_0116 = []

feed_ch4_36_0146 = []
v_36_0146 = []
temperature_36_0146 = []
pressure_36_0146 = []
experimental_36_0146 = []

feed_ch4_36_0129 = []
v_36_0129 = []
temperature_36_0129 = []
pressure_36_0129 = []
experimental_36_0129 = []

feed_ch4_36_022 = []
v_36_022 = []
temperature_36_022 = []
pressure_36_022 = []
experimental_36_022 = []

feed_ch4_36_0276 = []
v_36_0276 = []
temperature_36_0276 = []
pressure_36_0276 = []
experimental_36_0276 = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    # verificar o valor da coluna B (index 1) e fazer o append nas listas correspondentes
    if row[0] == 0.043:
        feed_ch4_36_0043.append(row[0])
        v_36_0043.append(row[3])
        temperature_36_0043.append(row[1])
        pressure_36_0043.append(row[2])
        experimental_36_0043.append(row[4])
    elif row[0] == 0.091:
        feed_ch4_36_0091.append(row[0])
        v_36_0091.append(row[3])
        temperature_36_0091.append(row[1])
        pressure_36_0091.append(row[2])
        experimental_36_0091.append(row[4])
    elif row[0] == 0.116:
        feed_ch4_36_0116.append(row[1])
        v_36_0116.append(row[3])
        temperature_36_0116.append(row[1])
        pressure_36_0116.append(row[2])
        experimental_36_0116.append(row[4])
    elif row[0] == 0.146:
        feed_ch4_36_0146.append(row[1])
        v_36_0146.append(row[3])
        temperature_36_0146.append(row[1])
        pressure_36_0146.append(row[2])
        experimental_36_0146.append(row[4])
    elif row[0] == 0.129:
        feed_ch4_36_0129.append(row[1])
        v_36_0129.append(row[3])
        temperature_36_0129.append(row[1])
        pressure_36_0129.append(row[2])
        experimental_36_0129.append(row[4])
    elif row[0] == 0.22:
        feed_ch4_36_022.append(row[1])
        v_36_022.append(row[3])
        temperature_36_022.append(row[1])
        pressure_36_022.append(row[2])
        experimental_36_022.append(row[4])
    elif row[0] == 0.276:
        feed_ch4_36_0276.append(row[1])
        v_36_0276.append(row[3])
        temperature_36_0276.append(row[1])
        pressure_36_0276.append(row[2])
        experimental_36_0276.append(row[4])

# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Data Fraction Test\Fraction Liquid Test\fraction_36_liquid_test.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [
    (0.043, (feed_ch4_36_0043, v_36_0043, temperature_36_0043, pressure_36_0043, experimental_36_0043)),
    (0.091, (feed_ch4_36_0091, v_36_0091, temperature_36_0091, pressure_36_0091, experimental_36_0091)),
    (0.116, (feed_ch4_36_0116, v_36_0116, temperature_36_0116, pressure_36_0116, experimental_36_0116)),
    (0.146, (feed_ch4_36_0146, v_36_0146, temperature_36_0146, pressure_36_0146, experimental_36_0146)),
    (0.129, (feed_ch4_36_0129, v_36_0129, temperature_36_0129, pressure_36_0129, experimental_36_0129)),
    (0.22, (feed_ch4_36_022, v_36_022, temperature_36_022, pressure_36_022, experimental_36_022)),
    (0.276, (feed_ch4_36_0276, v_36_0276, temperature_36_0276, pressure_36_0276, experimental_36_0276))]:
    sheet_name = f'frac_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CH4 Liquid {ref_value}', f'V Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Rho Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)