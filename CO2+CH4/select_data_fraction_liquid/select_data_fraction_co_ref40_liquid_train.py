import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Filtered References Train\References Liquid Train\filtered_data_references_liquid_train.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - NOVOS VALORES:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[1]

# inicializar listas para armazenar os dados:
feed_ch4_40_00010 = []
feed_co2_00010 = []
temperature_40_00010 = []
pressure_40_00010 = []
experimental_40_00010 = []

feed_ch4_40_00000 = []
feed_co2_00000 = []
temperature_40_00000 = []
pressure_40_00000 = []
experimental_40_00000 = []

feed_ch4_40_00030 = []
feed_co2_00030 = []
temperature_40_00030 = []
pressure_40_00030 = []
experimental_40_00030 = []

feed_ch4_40_00040 = []
feed_co2_00040 = []
temperature_40_00040 = []
pressure_40_00040 = []
experimental_40_00040 = []

feed_ch4_40_00070 = []
feed_co2_00070 = []
temperature_40_00070 = []
pressure_40_00070 = []
experimental_40_00070 = []

feed_ch4_40_00110 = []
feed_co2_00110 = []
temperature_40_00110 = []
pressure_40_00110 = []
experimental_40_00110 = []

feed_ch4_40_00160 = []
feed_co2_00160 = []
temperature_40_00160 = []
pressure_40_00160 = []
experimental_40_00160 = []

feed_ch4_40_000130 = []
feed_co2_000130 = []
temperature_40_000130 = []
pressure_40_000130 = []
experimental_40_000130 = []

feed_ch4_40_00006 = []
feed_co2_00006 = []
temperature_40_00006 = []
pressure_40_00006 = []
experimental_40_00006 = []

feed_ch4_40_00004 = []
feed_co2_00004 = []
temperature_40_00004 = []
pressure_40_00004 = []
experimental_40_00004 = []

feed_ch4_40_00001 = []
feed_co2_00001 = []
temperature_40_00001 = []
pressure_40_00001 = []
experimental_40_00001 = []

feed_ch4_40_00009 = []
feed_co2_00009 = []
temperature_40_00009 = []
pressure_40_00009 = []
experimental_40_00009 = []

feed_ch4_40_00014 = []
feed_co2_00014 = []
temperature_40_00014 = []
pressure_40_00014 = []
experimental_40_00014 = []

feed_ch4_40_00024 = []
feed_co2_00024 = []
temperature_40_00024 = []
pressure_40_00024 = []
experimental_40_00024 = []

feed_ch4_40_00041 = []
feed_co2_00041 = []
temperature_40_00041 = []
pressure_40_00041 = []
experimental_40_00041 = []

feed_ch4_40_00063 = []
feed_co2_00063 = []
temperature_40_00063 = []
pressure_40_00063 = []
experimental_40_00063 = []

feed_ch4_40_00084 = []
feed_co2_00084 = []
temperature_40_00084 = []
pressure_40_00084 = []
experimental_40_00084 = []

feed_ch4_40_00142 = []
feed_co2_00142 = []
temperature_40_00142 = []
pressure_40_00142 = []
experimental_40_00142 = []

feed_ch4_40_00020 = []
feed_co2_00020 = []
temperature_40_00020 = []
pressure_40_00020 = []
experimental_40_00020 = []

feed_ch4_40_00200 = []
feed_co2_00200 = []
temperature_40_00200 = []
pressure_40_00200 = []
experimental_40_00200 = []

feed_ch4_40_00250 = []
feed_co2_00250 = []
temperature_40_00250 = []
pressure_40_00250 = []
experimental_40_00250 = []

feed_ch4_40_00011 = []
feed_co2_00011 = []
temperature_40_00011 = []
pressure_40_00011 = []
experimental_40_00011 = []

feed_ch4_40_00007 = []
feed_co2_00007 = []
temperature_40_00007 = []
pressure_40_00007 = []
experimental_40_00007 = []

feed_ch4_40_00033 = []
feed_co2_00033 = []
temperature_40_00033 = []
pressure_40_00033 = []
experimental_40_00033 = []

feed_ch4_40_00071 = []
feed_co2_00071 = []
temperature_40_00071 = []
pressure_40_00071 = []
experimental_40_00071 = []

feed_ch4_40_00108 = []
feed_co2_00108 = []
temperature_40_00108 = []
pressure_40_00108 = []
experimental_40_00108 = []

feed_ch4_40_00151 = []
feed_co2_00151 = []
temperature_40_00151 = []
pressure_40_00151 = []
experimental_40_00151 = []

feed_ch4_40_00198 = []
feed_co2_00198 = []
temperature_40_00198 = []
pressure_40_00198 = []
experimental_40_00198 = []

feed_ch4_40_00171 = []
feed_co2_00171 = []
temperature_40_00171 = []
pressure_40_00171 = []
experimental_40_00171 = []

# Adicione todas as listas necessárias...

# Preencher as listas com os dados
for row in sheet.iter_rows(min_row=2, values_only=True):  # Supondo que 'data' seja a lista de dados

    if row[0] == 0.001:
        feed_ch4_40_00010.append(row[0])
        feed_co2_00010.append(row[5])
        temperature_40_00010.append(row[1])
        pressure_40_00010.append(row[2])
        experimental_40_00010.append(row[4])

    elif row[0] == 0.0000:
        feed_ch4_40_00000.append(row[0])
        feed_co2_00000.append(row[5])
        temperature_40_00000.append(row[1])
        pressure_40_00000.append(row[2])
        experimental_40_00000.append(row[4])

    elif row[0] == 0.003:
        feed_ch4_40_00030.append(row[0])
        feed_co2_00030.append(row[5])
        temperature_40_00030.append(row[1])
        pressure_40_00030.append(row[2])
        experimental_40_00030.append(row[4])

    elif row[0] == 0.004:
        feed_ch4_40_00040.append(row[0])
        feed_co2_00040.append(row[5])
        temperature_40_00040.append(row[1])
        pressure_40_00040.append(row[2])
        experimental_40_00040.append(row[4])

    elif row[0] == 0.007:
        feed_ch4_40_00070.append(row[0])
        feed_co2_00070.append(row[5])
        temperature_40_00070.append(row[1])
        pressure_40_00070.append(row[2])
        experimental_40_00070.append(row[4])

    elif row[0] == 0.011:
        feed_ch4_40_00110.append(row[0])
        feed_co2_00110.append(row[5])
        temperature_40_00110.append(row[1])
        pressure_40_00110.append(row[2])
        experimental_40_00110.append(row[4])

    elif row[0] == 0.016:
        feed_ch4_40_00160.append(row[0])
        feed_co2_00160.append(row[5])
        temperature_40_00160.append(row[1])
        pressure_40_00160.append(row[2])
        experimental_40_00160.append(row[4])

    elif row[0] == 0.0013:
        feed_ch4_40_000130.append(row[0])
        feed_co2_000130.append(row[5])
        temperature_40_000130.append(row[1])
        pressure_40_000130.append(row[2])
        experimental_40_000130.append(row[4])

    elif row[0] == 0.0006:
        feed_ch4_40_00006.append(row[0])
        feed_co2_00006.append(row[5])
        temperature_40_00006.append(row[1])
        pressure_40_00006.append(row[2])
        experimental_40_00006.append(row[4])

    elif row[0] == 0.0004:
        feed_ch4_40_00004.append(row[0])
        feed_co2_00004.append(row[5])
        temperature_40_00004.append(row[1])
        pressure_40_00004.append(row[2])
        experimental_40_00004.append(row[4])

    elif row[0] == 0.0001:
        feed_ch4_40_00001.append(row[0])
        feed_co2_00001.append(row[5])
        temperature_40_00001.append(row[1])
        pressure_40_00001.append(row[2])
        experimental_40_00001.append(row[4])

    elif row[0] == 0.0009:
        feed_ch4_40_00009.append(row[0])
        feed_co2_00009.append(row[5])
        temperature_40_00009.append(row[1])
        pressure_40_00009.append(row[2])
        experimental_40_00009.append(row[4])

    elif row[0] == 0.0014:
        feed_ch4_40_00014.append(row[0])
        feed_co2_00014.append(row[5])
        temperature_40_00014.append(row[1])
        pressure_40_00014.append(row[2])
        experimental_40_00014.append(row[4])

    elif row[0] == 0.0024:
        feed_ch4_40_00024.append(row[0])
        feed_co2_00024.append(row[5])
        temperature_40_00024.append(row[1])
        pressure_40_00024.append(row[2])
        experimental_40_00024.append(row[4])

    elif row[0] == 0.0041:
        feed_ch4_40_00041.append(row[0])
        feed_co2_00041.append(row[5])
        temperature_40_00041.append(row[1])
        pressure_40_00041.append(row[2])
        experimental_40_00041.append(row[4])

    elif row[0] == 0.0063:
        feed_ch4_40_00063.append(row[0])
        feed_co2_00063.append(row[5])
        temperature_40_00063.append(row[1])
        pressure_40_00063.append(row[2])
        experimental_40_00063.append(row[4])

    elif row[0] == 0.0084:
        feed_ch4_40_00084.append(row[0])
        feed_co2_00084.append(row[5])
        temperature_40_00084.append(row[1])
        pressure_40_00084.append(row[2])
        experimental_40_00084.append(row[4])

    elif row[0] == 0.0142:
        feed_ch4_40_00142.append(row[0])
        feed_co2_00142.append(row[5])
        temperature_40_00142.append(row[1])
        pressure_40_00142.append(row[2])
        experimental_40_00142.append(row[4])

    elif row[0] == 0.002:
        feed_ch4_40_00020.append(row[0])
        feed_co2_00020.append(row[5])
        temperature_40_00020.append(row[1])
        pressure_40_00020.append(row[2])
        experimental_40_00020.append(row[4])

    elif row[0] == 0.02:
        feed_ch4_40_00200.append(row[0])
        feed_co2_00200.append(row[5])
        temperature_40_00200.append(row[1])
        pressure_40_00200.append(row[2])
        experimental_40_00200.append(row[4])

    elif row[0] == 0.025:
        feed_ch4_40_00250.append(row[0])
        feed_co2_00250.append(row[5])
        temperature_40_00250.append(row[1])
        pressure_40_00250.append(row[2])
        experimental_40_00250.append(row[4])

    elif row[0] == 0.0011:
        feed_ch4_40_00011.append(row[0])
        feed_co2_00011.append(row[5])
        temperature_40_00011.append(row[1])
        pressure_40_00011.append(row[2])
        experimental_40_00011.append(row[4])

    elif row[0] == 0.0007:
        feed_ch4_40_00007.append(row[0])
        feed_co2_00007.append(row[5])
        temperature_40_00007.append(row[1])
        pressure_40_00007.append(row[2])
        experimental_40_00007.append(row[4])

    elif row[0] == 0.0033:
        feed_ch4_40_00033.append(row[0])
        feed_co2_00033.append(row[5])
        temperature_40_00033.append(row[1])
        pressure_40_00033.append(row[2])
        experimental_40_00033.append(row[4])

    elif row[0] == 0.0071:
        feed_ch4_40_00071.append(row[0])
        feed_co2_00071.append(row[5])
        temperature_40_00071.append(row[1])
        pressure_40_00071.append(row[2])
        experimental_40_00071.append(row[4])

    elif row[0] == 0.0108:
        feed_ch4_40_00108.append(row[0])
        feed_co2_00108.append(row[5])
        temperature_40_00108.append(row[1])
        pressure_40_00108.append(row[2])
        experimental_40_00108.append(row[4])

    elif row[0] == 0.0151:
        feed_ch4_40_00151.append(row[0])
        feed_co2_00151.append(row[5])
        temperature_40_00151.append(row[1])
        pressure_40_00151.append(row[2])
        experimental_40_00151.append(row[4])

    elif row[0] == 0.0198:
        feed_ch4_40_00198.append(row[0])
        feed_co2_00198.append(row[5])
        temperature_40_00198.append(row[1])
        pressure_40_00198.append(row[2])
        experimental_40_00198.append(row[4])

    elif row[0] == 0.0171:
        feed_ch4_40_00171.append(row[0])
        feed_co2_00171.append(row[5])
        temperature_40_00171.append(row[1])
        pressure_40_00171.append(row[2])
        experimental_40_00171.append(row[4])

# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Data Fraction Train\Fraction Liquid Train\fraction_40_liquid_train.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [
    (0.0010, (feed_ch4_40_00010, feed_co2_00010, temperature_40_00010, pressure_40_00010, experimental_40_00010)),
    (0.0000, (feed_ch4_40_00000, feed_co2_00000, temperature_40_00000, pressure_40_00000, experimental_40_00000)),
    (0.0030, (feed_ch4_40_00030, feed_co2_00030, temperature_40_00030, pressure_40_00030, experimental_40_00030)),
    (0.0040, (feed_ch4_40_00040, feed_co2_00040, temperature_40_00040, pressure_40_00040, experimental_40_00040)),
    (0.0070, (feed_ch4_40_00070, feed_co2_00070, temperature_40_00070, pressure_40_00070, experimental_40_00070)),
    (0.0110, (feed_ch4_40_00110, feed_co2_00110, temperature_40_00110, pressure_40_00110, experimental_40_00110)),
    (0.0160, (feed_ch4_40_00160, feed_co2_00160, temperature_40_00160, pressure_40_00160, experimental_40_00160)),
    (0.0013, (feed_ch4_40_000130, feed_co2_000130, temperature_40_000130, pressure_40_000130, experimental_40_000130)),
    (0.0006, (feed_ch4_40_00006, feed_co2_00006, temperature_40_00006, pressure_40_00006, experimental_40_00006)),
    (0.0004, (feed_ch4_40_00004, feed_co2_00004, temperature_40_00004, pressure_40_00004, experimental_40_00004)),
    (0.0001, (feed_ch4_40_00001, feed_co2_00001, temperature_40_00001, pressure_40_00001, experimental_40_00001)),
    (0.0009, (feed_ch4_40_00009, feed_co2_00009, temperature_40_00009, pressure_40_00009, experimental_40_00009)),
    (0.0014, (feed_ch4_40_00014, feed_co2_00014, temperature_40_00014, pressure_40_00014, experimental_40_00014)),
    (0.0024, (feed_ch4_40_00024, feed_co2_00024, temperature_40_00024, pressure_40_00024, experimental_40_00024)),
    (0.0041, (feed_ch4_40_00041, feed_co2_00041, temperature_40_00041, pressure_40_00041, experimental_40_00041)),
    (0.0063, (feed_ch4_40_00063, feed_co2_00063, temperature_40_00063, pressure_40_00063, experimental_40_00063)),
    (0.0084, (feed_ch4_40_00084, feed_co2_00084, temperature_40_00084, pressure_40_00084, experimental_40_00084)),
    (0.0142, (feed_ch4_40_00142, feed_co2_00142, temperature_40_00142, pressure_40_00142, experimental_40_00142)),
    (0.0020, (feed_ch4_40_00020, feed_co2_00020, temperature_40_00020, pressure_40_00020, experimental_40_00020)),
    (0.0200, (feed_ch4_40_00200, feed_co2_00200, temperature_40_00200, pressure_40_00200, experimental_40_00200)),
    (0.0250, (feed_ch4_40_00250, feed_co2_00250, temperature_40_00250, pressure_40_00250, experimental_40_00250)),
    (0.0011, (feed_ch4_40_00011, feed_co2_00011, temperature_40_00011, pressure_40_00011, experimental_40_00011)),
    (0.0007, (feed_ch4_40_00007, feed_co2_00007, temperature_40_00007, pressure_40_00007, experimental_40_00007)),
    (0.0033, (feed_ch4_40_00033, feed_co2_00033, temperature_40_00033, pressure_40_00033, experimental_40_00033)),
    (0.0071, (feed_ch4_40_00071, feed_co2_00071, temperature_40_00071, pressure_40_00071, experimental_40_00071)),
    (0.0108, (feed_ch4_40_00108, feed_co2_00108, temperature_40_00108, pressure_40_00108, experimental_40_00108)),
    (0.0151, (feed_ch4_40_00151, feed_co2_00151, temperature_40_00151, pressure_40_00151, experimental_40_00151)),
    (0.0198, (feed_ch4_40_00198, feed_co2_00198, temperature_40_00198, pressure_40_00198, experimental_40_00198)),
    (0.0171, (feed_ch4_40_00171, feed_co2_00171, temperature_40_00171, pressure_40_00171, experimental_40_00171)),
]:

    sheet_name = f'frac_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CH4 Liquid {ref_value}',  f'Feed CO2 Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Rho Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)