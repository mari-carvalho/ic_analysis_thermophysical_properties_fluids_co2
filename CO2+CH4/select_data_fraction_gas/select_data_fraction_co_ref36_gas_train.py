import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Filtered References Train\References Gas Train\filtered_data_references_gas_train.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - REFERÊNCIA 27 - 0.004:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
feed_ch4_36_058 = []
v_36_058 = []
temperature_36_058 = []
pressure_36_058 = []
experimental_36_058 = []

feed_ch4_36_0098 = []
v_36_0098 = []
temperature_36_0098 = []
pressure_36_0098 = []
experimental_36_0098 = []

feed_ch4_36_0137 = []
v_36_0137 = []
temperature_36_0137 = []
pressure_36_0137 = []
experimental_36_0137 = []

feed_ch4_36_0167 = []
v_36_0167 = []
temperature_36_0167 = []
pressure_36_0167 = []
experimental_36_0167 = []

feed_ch4_36_0177 = []
v_36_0177 = []
temperature_36_0177 = []
pressure_36_0177 = []
experimental_36_0177 = []

feed_ch4_36_0182 = []
v_36_0182 = []
temperature_36_0182 = []
pressure_36_0182 = []
experimental_36_0182 = []

feed_ch4_36_0247 = []
v_36_0247 = []
temperature_36_0247 = []
pressure_36_0247 = []
experimental_36_0247 = []

feed_ch4_36_03 = []
v_36_03 = []
temperature_36_03 = []
pressure_36_03 = []
experimental_36_03 = []

feed_ch4_36_0214 = []
v_36_0214 = []
temperature_36_0214 = []
pressure_36_0214 = []
experimental_36_0214 = []

feed_ch4_36_0301 = []
v_36_0301 = []
temperature_36_0301 = []
pressure_36_0301 = []
experimental_36_0301 = []

feed_ch4_36_0321 = []
v_36_0321 = []
temperature_36_0321 = []
pressure_36_0321 = []
experimental_36_0321 = []

feed_ch4_36_034 = []
v_36_034 = []
temperature_36_034 = []
pressure_36_034 = []
experimental_36_034 = []

feed_ch4_36_0349 = []
v_36_0349 = []
temperature_36_0349 = []
pressure_36_0349 = []
experimental_36_0349 = []

feed_ch4_36_0408 = []
v_36_0408 = []
temperature_36_0408 = []
pressure_36_0408 = []
experimental_36_0408 = []

feed_ch4_36_0508 = []
v_36_0508 = []
temperature_36_0508 = []
pressure_36_0508 = []
experimental_36_0508 = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    # verificar o valor da coluna B (index 1) e fazer o append nas listas correspondentes
    if row[0] == 0.58:
        feed_ch4_36_058.append(row[0])
        v_36_058.append(row[3])
        temperature_36_058.append(row[1])
        pressure_36_058.append(row[2])
        experimental_36_058.append(row[4])
    elif row[0] == 0.098:
        feed_ch4_36_0098.append(row[0])
        v_36_0098.append(row[3])
        temperature_36_0098.append(row[1])
        pressure_36_0098.append(row[2])
        experimental_36_0098.append(row[4])
    elif row[0] == 0.137:
        feed_ch4_36_0137.append(row[1])
        v_36_0137.append(row[3])
        temperature_36_0137.append(row[1])
        pressure_36_0137.append(row[2])
        experimental_36_0137.append(row[4])
    elif row[0] == 0.167:
        feed_ch4_36_0167.append(row[1])
        v_36_0167.append(row[3])
        temperature_36_0167.append(row[1])
        pressure_36_0167.append(row[2])
        experimental_36_0167.append(row[4])
    elif row[0] == 0.177:
        feed_ch4_36_0177.append(row[1])
        v_36_0177.append(row[3])
        temperature_36_0177.append(row[1])
        pressure_36_0177.append(row[2])
        experimental_36_0177.append(row[4])
    elif row[0] == 0.182:
        feed_ch4_36_0182.append(row[1])
        v_36_0182.append(row[3])
        temperature_36_0182.append(row[1])
        pressure_36_0182.append(row[2])
        experimental_36_0182.append(row[4])
    elif row[0] == 0.247:
        feed_ch4_36_0247.append(row[1])
        v_36_0247.append(row[3])
        temperature_36_0247.append(row[1])
        pressure_36_0247.append(row[2])
        experimental_36_0247.append(row[4])
    elif row[0] == 0.3:
        feed_ch4_36_03.append(row[1])
        v_36_03.append(row[3])
        temperature_36_03.append(row[1])
        pressure_36_03.append(row[2])
        experimental_36_03.append(row[4])
    elif row[0] == 0.214:
        feed_ch4_36_0214.append(row[0])
        v_36_0214.append(row[3])
        temperature_36_0214.append(row[1])
        pressure_36_0214.append(row[2])
        experimental_36_0214.append(row[4])
    elif row[0] == 0.301:
        feed_ch4_36_0301.append(row[0])
        v_36_0301.append(row[3])
        temperature_36_0301.append(row[1])
        pressure_36_0301.append(row[2])
        experimental_36_0301.append(row[4])
    elif row[0] == 0.321:
        feed_ch4_36_0321.append(row[0])
        v_36_0321.append(row[3])
        temperature_36_0321.append(row[1])
        pressure_36_0321.append(row[2])
        experimental_36_0321.append(row[4])
    elif row[0] == 0.34:
        feed_ch4_36_034.append(row[0])
        v_36_034.append(row[3])
        temperature_36_034.append(row[1])
        pressure_36_034.append(row[2])
        experimental_36_034.append(row[4])
    elif row[0] == 0.349:
        feed_ch4_36_0349.append(row[0])
        v_36_0349.append(row[3])
        temperature_36_0349.append(row[1])
        pressure_36_0349.append(row[2])
        experimental_36_0349.append(row[4])
    elif row[0] == 0.408:
        feed_ch4_36_0408.append(row[0])
        v_36_0408.append(row[3])
        temperature_36_0408.append(row[1])
        pressure_36_0408.append(row[2])
        experimental_36_0408.append(row[4])
    elif row[0] == 0.508:
        feed_ch4_36_0508.append(row[0])
        v_36_0508.append(row[3])
        temperature_36_0508.append(row[1])
        pressure_36_0508.append(row[1])
        experimental_36_0508.append(row[4])

# Exibindo os dados armazenados para cada conjunto:
print("Dados armazenados para 0.058:")
print("Feed CH4:", feed_ch4_36_058)
print("V:", v_36_058)
print("Temperature:", temperature_36_058)
print("Pressure:", pressure_36_058)
print("Experimental:", experimental_36_058)


# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Data Fraction Train\Fraction Gas Train\fraction_36_gas_train.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [
    (0.58, (feed_ch4_36_058, v_36_058, temperature_36_058, pressure_36_058, experimental_36_058)),
    (0.098, (feed_ch4_36_0098, v_36_0098, temperature_36_0098, pressure_36_0098, experimental_36_0098)),
    (0.137, (feed_ch4_36_0137, v_36_0137, temperature_36_0137, pressure_36_0137, experimental_36_0137)),
    (0.167, (feed_ch4_36_0167, v_36_0167, temperature_36_0167, pressure_36_0167, experimental_36_0167)),
    (0.177, (feed_ch4_36_0177, v_36_0177, temperature_36_0177, pressure_36_0177, experimental_36_0177)),
    (0.182, (feed_ch4_36_0182, v_36_0182, temperature_36_0182, pressure_36_0182, experimental_36_0182)),
    (0.247, (feed_ch4_36_0247, v_36_0247, temperature_36_0247, pressure_36_0247, experimental_36_0247)),
    (0.3, (feed_ch4_36_03, v_36_03, temperature_36_03, pressure_36_03, experimental_36_03)),
    (0.214, (feed_ch4_36_0214, v_36_0214, temperature_36_0214, pressure_36_0214, experimental_36_0214)),
    (0.301, (feed_ch4_36_0301, v_36_0301, temperature_36_0301, pressure_36_0301, experimental_36_0301)),
    (0.321, (feed_ch4_36_0321, v_36_0321, temperature_36_0321, pressure_36_0321, experimental_36_0321)),
    (0.34, (feed_ch4_36_034, v_36_034, temperature_36_034, pressure_36_034, experimental_36_034)),
    (0.349, (feed_ch4_36_0349, v_36_0349, temperature_36_0349, pressure_36_0349, experimental_36_0349)),
    (0.408, (feed_ch4_36_0408, v_36_0408, temperature_36_0408, pressure_36_0408, experimental_36_0408)),
    (0.508, (feed_ch4_36_0508, v_36_0508, temperature_36_0508, pressure_36_0508, experimental_36_0508))
]:
    sheet_name = f'frac_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CH4 gas {ref_value}', f'V Gas {ref_value}', f'Temperature Gas {ref_value}', f'Pressure Gas {ref_value}', f'Rho Gas {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)