import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Train_Test Liquid\dados_treinamento_com_densidade_liquid.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
# referência 36
feed_ch4_36 = []
feed_co_36 = []
v_36 = []
temperature_36 = []
pressure_36 = []
experimental_36 = []

feed_ch4_40 = []
feed_co_40 = []
v_40 = []
temperature_40 = []
pressure_40 = []
experimental_40 = []


# iterar pelas linhas da planilha a partir da segunda linha (ignorando o cabeçalho):
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7, values_only=True):
    # verificar se o valor da coluna G (index 5) é 36
    if row[0] == 36:
        feed_ch4_36.append(row[1])
        temperature_36.append(row[2])
        pressure_36.append(row[3])
        v_36.append(row[4])
        experimental_36.append(row[5])
        feed_co_36.append(row[6])
    elif row[0] == 40:
        feed_ch4_40.append(row[1])
        temperature_40.append(row[2])
        pressure_40.append(row[3])
        v_40.append(row[4])
        experimental_40.append(row[5])
        feed_co_40.append(row[6])

# Exibindo os dados armazenados:
print("CH4 Feed (Liquid 36):", feed_ch4_36)
print("V (Liquid 36):", v_36)
print("Temperature (Liquid 36):", temperature_36)
print("Pressure (Liquid 36):", pressure_36)
print("Experimental (Liquid 36):", experimental_36)
print("CO Feed (Liquid 36):", experimental_36)

print("CH4 Feed (Liquid 40):", feed_ch4_40)
print("V (Liquid 40):", v_40)
print("Temperature (Liquid 40):", temperature_40)
print("Pressure (Liquid 40):", pressure_40)
print("Experimental (Liquid 40):", experimental_40)
print("CO Feed (Liquid 40):", experimental_40)

# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CH4+CO2\Seleção Treino_Teste\Filtered References Train\References Liquid Train\filtered_data_references_liquid_train.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(36, (feed_ch4_36, v_36, temperature_36, pressure_36, experimental_36, feed_co_36)),
                        (40, (feed_ch4_40, v_40, temperature_40, pressure_40, experimental_40, feed_co_40))]:

    sheet_name = f'ref_{ref_value}_data_Liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CH4 Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'V Liquid {ref_value}', f'Rho Liquid {ref_value}', f'Feed CO2 Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)