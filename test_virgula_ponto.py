import numpy as np
import pandas as pd
import openpyxl

file_path = r"C:\Users\03950025081\Downloads\Metan Databank.xlsx"

workbook = openpyxl.load_workbook(file_path, data_only=True)
for sheet in workbook.sheetnames: # vai iterar sobre as planilhas (abas)
    worksheet = workbook[sheet] # vai selecionar a aba da vez
    for col in worksheet.iter_cols(): # vai iterar sobre todas as colunas da aba da vez
        for cell in col: # vai iterar em cada célula da coluna
            if isinstance(cell.value, str): # verifica se é uma string
                cell.value = str(cell.value).replace('.', ',') # converter o número para string e subtituir o ponto por vírgula


# salvar o arquivo modificado
novo_caminho = r"C:\Users\03950025081\Downloads\Metan Databank Modificado.xlsx"
workbook.save(novo_caminho)
