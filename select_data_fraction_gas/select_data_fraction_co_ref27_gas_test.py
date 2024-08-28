import numpy as np
import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\Filtered References Test\References Gas Test\filtered_data_references_gas_test.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO - REFERÊNCIA 27 - 0.004:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
feed_co2_27_004 = []
feed_co_27_004 = []
temperature_27_004 = []
pressure_27_004 = []
experimental_27_004 = []

feed_co2_27_007 = []
feed_co_27_007 = []
temperature_27_007 = []
pressure_27_007 = []
experimental_27_007 = []

feed_co2_27_0098 = []
feed_co_27_0098 = []
temperature_27_0098 = []
pressure_27_0098 = []
experimental_27_0098 = []

feed_co2_27_010107 = []
feed_co_27_010107 = []
temperature_27_010107 = []
pressure_27_010107 = []
experimental_27_010107 = []

feed_co2_27_05031 = []
feed_co_27_05031 = []
temperature_27_05031 = []
pressure_27_05031 = []
experimental_27_05031 = []

feed_co2_27_03 = []
feed_co_27_03 = []
temperature_27_03 = []
pressure_27_03 = []
experimental_27_03 = []

feed_co2_27_25254 = []
feed_co_27_25254 = []
temperature_27_25254 = []
pressure_27_25254 = []
experimental_27_25254 = []

feed_co2_27_065 = []
feed_co_27_065 = []
temperature_27_065 = []
pressure_27_065 = []
experimental_27_065 = []

feed_co2_27_50183 = []
feed_co_27_50183 = []
temperature_27_50183 = []
pressure_27_50183 = []
experimental_27_50183 = []

feed_co2_27_138 = []
feed_co_27_138 = []
temperature_27_138 = []
pressure_27_138 = []
experimental_27_138 = []

feed_co2_27_134 = []
feed_co_27_134 = []
temperature_27_134 = []
pressure_27_134 = []
experimental_27_134 = []

feed_co2_27_117 = []
feed_co_27_117 = []
temperature_27_117 = []
pressure_27_117 = []
experimental_27_117 = []

feed_co2_27_115 = []
feed_co_27_115 = []
temperature_27_115 = []
pressure_27_115 = []
experimental_27_115 = []

feed_co2_27_089 = []
feed_co_27_089 = []
temperature_27_089 = []
pressure_27_089 = []
experimental_27_089 = []

feed_co2_27_046 = []
feed_co_27_046 = []
temperature_27_046 = []
pressure_27_046 = []
experimental_27_046 = []

feed_co2_27_019 = []
feed_co_27_019 = []
temperature_27_019 = []
pressure_27_019 = []
experimental_27_019 = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    # verificar se o valor da coluna G (index 5) é 27
    if row[1] == 0.004:
        feed_co2_27_004.append(row[0])
        feed_co_27_004.append(row[1])
        temperature_27_004.append(row[2])
        pressure_27_004.append(row[3])
        experimental_27_004.append(row[4])
    elif row[1] == 0.007:
        feed_co2_27_007.append(row[0])
        feed_co_27_007.append(row[1])
        temperature_27_007.append(row[2])
        pressure_27_007.append(row[3])
        experimental_27_007.append(row[4])
    elif row[1] == 0.0098:
        feed_co2_27_0098.append(row[0])
        feed_co_27_0098.append(row[1])
        temperature_27_0098.append(row[2])
        pressure_27_0098.append(row[3])
        experimental_27_0098.append(row[4])
    elif row[1] == 0.010107:
        feed_co2_27_010107.append(row[0])
        feed_co_27_010107.append(row[1])
        temperature_27_010107.append(row[2])
        pressure_27_010107.append(row[3])
        experimental_27_010107.append(row[4])
    elif row[1] == 0.05031:
        feed_co2_27_05031.append(row[0])
        feed_co_27_05031.append(row[1])
        temperature_27_05031.append(row[2])
        pressure_27_05031.append(row[3])
        experimental_27_05031.append(row[4])
    elif row[1] == 0.03:
        feed_co2_27_03.append(row[0])
        feed_co_27_03.append(row[1])
        temperature_27_03.append(row[2])
        pressure_27_03.append(row[3])
        experimental_27_03.append(row[4])
    elif row[1] == 0.25254:
        feed_co2_27_25254.append(row[0])
        feed_co_27_25254.append(row[1])
        temperature_27_25254.append(row[2])
        pressure_27_25254.append(row[3])
        experimental_27_25254.append(row[4])
    elif row[1] == 0.065:
        feed_co2_27_065.append(row[0])
        feed_co_27_065.append(row[1])
        temperature_27_065.append(row[2])
        pressure_27_065.append(row[3])
        experimental_27_065.append(row[4])
    elif row[1] == 0.50183:
        feed_co2_27_50183.append(row[0])
        feed_co_27_50183.append(row[1])
        temperature_27_50183.append(row[2])
        pressure_27_50183.append(row[3])
        experimental_27_50183.append(row[4])
    elif row[1] == 0.138:
        feed_co2_27_138.append(row[0])
        feed_co_27_138.append(row[1])
        temperature_27_138.append(row[2])
        pressure_27_138.append(row[3])
        experimental_27_138.append(row[4])
    elif row[1] == 0.134:
        feed_co2_27_134.append(row[0])
        feed_co_27_134.append(row[1])
        temperature_27_134.append(row[2])
        pressure_27_134.append(row[3])
        experimental_27_134.append(row[4])
    elif row[1] == 0.117:
        feed_co2_27_117.append(row[0])
        feed_co_27_117.append(row[1])
        temperature_27_117.append(row[2])
        pressure_27_117.append(row[3])
        experimental_27_117.append(row[4])
    elif row[1] == 0.115:
        feed_co2_27_115.append(row[0])
        feed_co_27_115.append(row[1])
        temperature_27_115.append(row[2])
        pressure_27_115.append(row[3])
        experimental_27_115.append(row[4])
    elif row[1] == 0.089:
        feed_co2_27_089.append(row[0])
        feed_co_27_089.append(row[1])
        temperature_27_089.append(row[2])
        pressure_27_089.append(row[3])
        experimental_27_089.append(row[4])
    elif row[1] == 0.046:
        feed_co2_27_046.append(row[0])
        feed_co_27_046.append(row[1])
        temperature_27_046.append(row[2])
        pressure_27_046.append(row[3])
        experimental_27_046.append(row[4])
    elif row[1] == 0.019:
        feed_co2_27_019.append(row[0])
        feed_co_27_019.append(row[1])
        temperature_27_019.append(row[2])
        pressure_27_019.append(row[3])
        experimental_27_019.append(row[4])

# Exibindo os dados armazenados:

print("CO2 Feed (Liquid 27 - 0.004):", feed_co2_27_004)
print("CO Feed (Liquid 27 - 0.004):", feed_co_27_004)
print("Temperature (Liquid 27 - 0.004):", temperature_27_004)
print("Pressure (Liquid 27 - 0.004):", pressure_27_004)
print("Experimental (Liquid 27 - 0.004):", experimental_27_004)

print("CO2 Feed (Liquid 27 - 0.007):", feed_co2_27_007)
print("CO Feed (Liquid 27 - 0.007):", feed_co_27_007)
print("Temperature (Liquid 27 - 0.007):", temperature_27_007)
print("Pressure (Liquid 27 - 0.007):", pressure_27_007)
print("Experimental (Liquid 27 - 0.007):", experimental_27_007)

print("CO2 Feed (Liquid 27 - 0.0098):", feed_co2_27_0098)
print("CO Feed (Liquid 27 - 0.0098):", feed_co_27_0098)
print("Temperature (Liquid 27 - 0.0098):", temperature_27_0098)
print("Pressure (Liquid 27 - 0.0098):", pressure_27_0098)
print("Experimental (Liquid 27 - 0.0098):", experimental_27_0098)

print("CO2 Feed (Liquid 27 - 0.010107):", feed_co2_27_010107)
print("CO Feed (Liquid 27 - 0.010107):", feed_co_27_010107)
print("Temperature (Liquid 27 - 0.010107):", temperature_27_010107)
print("Pressure (Liquid 27 - 0.010107):", pressure_27_010107)
print("Experimental (Liquid 27 - 0.010107):", experimental_27_010107)

print("CO2 Feed (Liquid 27 - 0.05031):", feed_co2_27_05031)
print("CO Feed (Liquid 27 - 0.05031):", feed_co_27_05031)
print("Temperature (Liquid 27 - 0.05031):", temperature_27_05031)
print("Pressure (Liquid 27 - 0.05031):", pressure_27_05031)
print("Experimental (Liquid 27 - 0.05031):", experimental_27_05031)

print("CO2 Feed (Liquid 27 - 0.03):", feed_co2_27_03)
print("CO Feed (Liquid 27 - 0.03):", feed_co_27_03)
print("Temperature (Liquid 27 - 0.03):", temperature_27_03)
print("Pressure (Liquid 27 - 0.03):", pressure_27_03)
print("Experimental (Liquid 27 - 0.03):", experimental_27_03)

print("CO2 Feed (Liquid 27 - 0.25254):", feed_co2_27_25254)
print("CO Feed (Liquid 27 - 0.25254):", feed_co_27_25254)
print("Temperature (Liquid 27 - 0.25254):", temperature_27_25254)
print("Pressure (Liquid 27 - 0.25254):", pressure_27_25254)
print("Experimental (Liquid 27 - 0.25254):", experimental_27_25254)

print("CO2 Feed (Liquid 27 - 0.065):", feed_co2_27_065)
print("CO Feed (Liquid 27 - 0.065):", feed_co_27_065)
print("Temperature (Liquid 27 - 0.065):", temperature_27_065)
print("Pressure (Liquid 27 - 0.065):", pressure_27_065)
print("Experimental (Liquid 27 - 0.065):", experimental_27_065)

print("CO2 Feed (Liquid 27 - 0.50183):", feed_co2_27_50183)
print("CO Feed (Liquid 27 - 0.50183):", feed_co_27_50183)
print("Temperature (Liquid 27 - 0.50183):", temperature_27_50183)
print("Pressure (Liquid 27 - 0.50183):", pressure_27_50183)
print("Experimental (Liquid 27 - 0.50183):", experimental_27_50183)

print("CO2 Feed (Liquid 27 - 0.138):", feed_co2_27_138)
print("CO Feed (Liquid 27 - 0.138):", feed_co_27_138)
print("Temperature (Liquid 27 - 0.138):", temperature_27_138)
print("Pressure (Liquid 27 - 0.138):", pressure_27_138)
print("Experimental (Liquid 27 - 0.138):", experimental_27_138)

print("CO2 Feed (Liquid 27 - 0.134):", feed_co2_27_134)
print("CO Feed (Liquid 27 - 0.134):", feed_co_27_134)
print("Temperature (Liquid 27 - 0.134):", temperature_27_134)
print("Pressure (Liquid 27 - 0.134):", pressure_27_134)
print("Experimental (Liquid 27 - 0.134):", experimental_27_134)

print("CO2 Feed (Liquid 27 - 0.117):", feed_co2_27_117)
print("CO Feed (Liquid 27 - 0.117):", feed_co_27_117)
print("Temperature (Liquid 27 - 0.117):", temperature_27_117)
print("Pressure (Liquid 27 - 0.117):", pressure_27_117)
print("Experimental (Liquid 27 - 0.117):", experimental_27_117)

print("CO2 Feed (Liquid 27 - 0.115):", feed_co2_27_115)
print("CO Feed (Liquid 27 - 0.115):", feed_co_27_115)
print("Temperature (Liquid 27 - 0.115):", temperature_27_115)
print("Pressure (Liquid 27 - 0.115):", pressure_27_115)
print("Experimental (Liquid 27 - 0.115):", experimental_27_115)

print("CO2 Feed (Liquid 27 - 0.089):", feed_co2_27_089)
print("CO Feed (Liquid 27 - 0.089):", feed_co_27_089)
print("Temperature (Liquid 27 - 0.089):", temperature_27_089)
print("Pressure (Liquid 27 - 0.089):", pressure_27_089)
print("Experimental (Liquid 27 - 0.089):", experimental_27_089)

print("CO2 Feed (Liquid 27 - 0.046):", feed_co2_27_046)
print("CO Feed (Liquid 27 - 0.046):", feed_co_27_046)
print("Temperature (Liquid 27 - 0.046):", temperature_27_046)
print("Pressure (Liquid 27 - 0.046):", pressure_27_046)
print("Experimental (Liquid 27 - 0.046):", experimental_27_046)

print("CO2 Feed (Liquid 27 - 0.046):", feed_co2_27_019)
print("CO Feed (Liquid 27 - 0.046):", feed_co_27_019)
print("Temperature (Liquid 27 - 0.046):", temperature_27_019)
print("Pressure (Liquid 27 - 0.046):", pressure_27_019)
print("Experimental (Liquid 27 - 0.046):", experimental_27_019)

# criando a nova planilha:
# caminho do arquivo:
filtered_file_path = r'C:\Users\03950025081\Desktop\Simulações Thermobuilder\Seleção Treino_Teste\Data Fraction Test\Fraction Gas Test\fraction_27_gas_test.xlsx'

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(0.004, (feed_co2_27_004, feed_co_27_004, temperature_27_004, pressure_27_004, experimental_27_004)),
                        (0.007, (feed_co2_27_007, feed_co_27_007, temperature_27_007, pressure_27_007, experimental_27_007)),
                        (0.0098, (feed_co2_27_0098, feed_co_27_0098, temperature_27_0098, pressure_27_0098, experimental_27_0098)),
                        (0.010107, (feed_co2_27_010107, feed_co_27_010107, temperature_27_010107, pressure_27_010107, experimental_27_010107)),
                        (0.05031, (feed_co2_27_05031, feed_co_27_05031, temperature_27_05031, pressure_27_05031, experimental_27_05031)),
                        (0.03, (feed_co2_27_03, feed_co_27_03, temperature_27_03, pressure_27_03, experimental_27_03)),
                        (0.25254, (feed_co2_27_25254, feed_co_27_25254, temperature_27_25254, pressure_27_25254, experimental_27_25254)),
                        (0.065, (feed_co2_27_065, feed_co_27_065, temperature_27_065, pressure_27_065, experimental_27_065)),
                        (0.50183, (feed_co2_27_50183, feed_co_27_50183, temperature_27_50183, pressure_27_50183, experimental_27_50183)),
                        (0.138, (feed_co2_27_138, feed_co_27_138, temperature_27_138, pressure_27_138, experimental_27_138)),
                        (0.134, (feed_co2_27_134, feed_co_27_134, temperature_27_134, pressure_27_134, experimental_27_134)),
                        (0.117, (feed_co2_27_117, feed_co_27_117, temperature_27_117, pressure_27_117, experimental_27_117)),
                        (0.115, (feed_co2_27_115, feed_co_27_115, temperature_27_115, pressure_27_115, experimental_27_115)),
                        (0.089, (feed_co2_27_089, feed_co_27_089, temperature_27_089, pressure_27_089, experimental_27_089)),
                        (0.046, (feed_co2_27_046, feed_co_27_046, temperature_27_046, pressure_27_046, experimental_27_046)),
                        (0.019, (feed_co2_27_019, feed_co_27_019, temperature_27_019, pressure_27_019, experimental_27_019))]:

    sheet_name = f'frac_{ref_value}_data_liquid'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CO2 Liquid {ref_value}', f'Feed CO Liquid {ref_value}', f'Temperature Liquid {ref_value}', f'Pressure Liquid {ref_value}', f'Experimental Liquid {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)