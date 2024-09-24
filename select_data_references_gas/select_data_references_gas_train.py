import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# Caminho do arquivo
file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO+CO2\Seleção Treino_Teste\Train_Test Gas\dados_treinamento_com_densidade_gas.xlsx"
workbook = load_workbook(file_path)

# DADOS LÍQUIDO:
# selecionar a primeira aba da planilha:
sheet = workbook.worksheets[0]

# inicializar listas para armazenar os dados:
# referência 27
feed_co2_27 = []
feed_co_27 = []
temperature_27 = []
pressure_27 = []
experimental_27 = []

# referência 25
feed_co2_25 = []
feed_co_25 = []
temperature_25 = []
pressure_25 = []
experimental_25 = []

# referência 7
feed_co2_7 = []
feed_co_7 = []
temperature_7 = []
pressure_7 = []
experimental_7 = []

# iterar pelas linhas da planilha a partir da segunda linha (ignorando o cabeçalho):
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
    # verificar se o valor da coluna G (index 5) é 27
    if row[5] == 27:
        feed_co2_27.append(row[0])
        feed_co_27.append(row[1])
        temperature_27.append(row[2])
        pressure_27.append(row[3])
        experimental_27.append(row[4])
    elif row[5] == 25:
        feed_co2_25.append(row[0])
        feed_co_25.append(row[1])
        temperature_25.append(row[2])
        pressure_25.append(row[3])
        experimental_25.append(row[4])
    elif row[5] == 7:
        feed_co2_7.append(row[0])
        feed_co_7.append(row[1])
        temperature_7.append(row[2])
        pressure_7.append(row[3])
        experimental_7.append(row[4])

# Exibindo os dados armazenados:
print("CO2 Feed (gas 27):", feed_co2_27)
print("CO Feed (gas 27):", feed_co_27)
print("Temperature (gas 27):", temperature_27)
print("Pressure (gas 27):", pressure_27)
print("Experimental (gas 27):", experimental_27)

# Exibindo os dados armazenados:
print("CO2 Feed (gas 25):", feed_co2_25)
print("CO Feed (gas 25):", feed_co_25)
print("Temperature (gas 25):", temperature_25)
print("Pressure (gas 25):", pressure_25)
print("Experimental (gas 25):", experimental_25)

# Exibindo os dados armazenados:
print("CO2 Feed (gas 7):", feed_co2_7)
print("CO Feed (gas 7):", feed_co_7)
print("Temperature (gas 7):", temperature_7)
print("Pressure (gas 7):", pressure_7)
print("Experimental (gas 7):", experimental_7)

# caminho do arquivo:
filtered_file_path = r"C:\Users\03950025081\Desktop\Simulações Thermobuilder\CO+CO2\Seleção Treino_Teste\Filtered References Train\References Gas Train\filtered_data_references_gas_train.xlsx"

# carregar a planilha existente:
try:
    workbook = load_workbook(filtered_file_path)
except FileNotFoundError:
    # se o arquivo não existir, criar um novo Workbook:
    workbook = Workbook()

# excluir todas as bas existentes:
for sheet in workbook.sheetnames:
    del workbook[sheet]

# adicionar os dados filtrados do líquido:
for ref_value, data in [(27, (feed_co2_27, feed_co_27, temperature_27, pressure_27, experimental_27)),
                        (25, (feed_co2_25, feed_co_25, temperature_25, pressure_25, experimental_25)),
                        (7, (feed_co2_7, feed_co_7, temperature_7, pressure_7, experimental_7))]:

    sheet_name = f'ref_{ref_value}_data_gas'
    ws = workbook.create_sheet(title=sheet_name)

    headers = [f'Feed CO2 gas {ref_value}', f'Feed CO gas {ref_value}', f'Temperature gas {ref_value}', f'Pressure gas {ref_value}', f'Experimental gas {ref_value}']
    ws.append(headers)

    rows = zip(*data)
    for row in rows:
        ws.append(row)

# salvar a nova planilha:
workbook.save(filtered_file_path)